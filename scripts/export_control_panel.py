#!/usr/bin/env python3
"""
export_control_panel.py
-----------------------
Exporta artefatos estruturados do OP Control Panel:
  reports/trades_snapshot_YYYYMMDD.json  — estado atual de todos os trades
  reports/trade_history.parquet          — série temporal completa (PnL/Delta)
  reports/monthly_summary.csv            — resumo de trades fechados por mês

Uso:
  python scripts/export_control_panel.py [--xlsx path] [--gdrive-id SHEET_ID]

Google Drive (opcional):
  Coloque o arquivo .credentials/gdrive_credentials.json (OAuth2 do GCP).
  Na primeira execução abre o browser para autenticação; depois usa token salvo.
  Ou passe o ID via --gdrive-id ou variável de ambiente GDRIVE_FILE_ID.
"""

import argparse
import json
import os
import re
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

ROOT = Path(__file__).parent.parent
DEFAULT_XLSX = ROOT / "data" / "control_panel" / "OP Control Panel.xlsx"
REPORTS_DIR  = ROOT / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

CREDENTIALS_DIR  = ROOT / ".credentials"
GDRIVE_CREDS     = CREDENTIALS_DIR / "gdrive_credentials.json"
GDRIVE_TOKEN     = CREDENTIALS_DIR / "gdrive_token.json"
GDRIVE_SCOPES    = ["https://www.googleapis.com/auth/drive.readonly"]
NON_INTERACTIVE_ENV_VARS = ("RENDER", "RENDER_SERVICE_ID", "VERCEL", "CI")

# ─── Environment mapping ─────────────────────────────────────────────────────
ENV_MAP = {
    "Live":         "CZ_Live",
    "Live-apr":     "CZ_Live",
    "Forward JS":   "JS_Forward",
    "FOR Trades":   "CZ_Forward",
    "FOR JS":       "JS_Forward",
}

# Mapeia cada aba visual para seu ambiente normalizado
SHEET_ENV_MAP = {
    "APR26":        "CZ_Live",
    "MAR26":        "CZ_Live",
    "JS APR26":     "JS_Forward",
    "JS-FOR MAR26": "JS_Forward",
    "FOR Trades":   "CZ_Forward",
}

def normalize_env(env: str) -> str:
    return ENV_MAP.get(str(env).strip(), str(env).strip())

def normalize_name(name: str) -> str:
    if not name:
        return ""
    return re.sub(r"\s+", " ", str(name).replace(" - CLOSED", "")).strip()


# ─── Loaders ─────────────────────────────────────────────────────────────────

def _parse_money(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if not (isinstance(val, float) and np.isnan(val)) else None
    s = str(val).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _parse_dte(val) -> tuple[str | None, int | None]:
    """Parse DTE cell preserving raw string AND extracting the first integer.

    João writes calendar-spread DTE as "7 / 10", "14/21" (front/back leg DTEs)
    or as a plain number "7" for single-leg trades. _parse_money() drops the
    spread variants — we keep both forms so the UI shows the original string
    and downstream math (dte_remaining etc) gets the front-leg int.
    """
    if val is None:
        return None, None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and np.isnan(val):
            return None, None
        return str(int(val)) if float(val).is_integer() else str(val), int(val)
    raw = str(val).strip()
    if not raw:
        return None, None
    m = re.search(r"\d+", raw)
    return raw, int(m.group()) if m else None


def _parse_sheet_date(val) -> date | None:
    """Coerce Excel/Sheets date cells, formatted strings, or serials to date."""
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        if 20000 <= float(val) <= 80000:
            return (datetime(1899, 12, 30) + timedelta(days=float(val))).date()
        return None
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        dot_match = re.match(r"^(\d{1,2})\.(\d{1,2})(?:\.(\d{2,4}))?$", s)
        if dot_match:
            day = int(dot_match.group(1))
            month = int(dot_match.group(2))
            year_raw = dot_match.group(3)
            year = int(year_raw) if year_raw else date.today().year
            if year < 100:
                year += 2000
            try:
                return date(year, month, day)
            except ValueError:
                return None
        parsed = pd.to_datetime(s, errors="coerce", dayfirst=True)
        if pd.notna(parsed):
            return parsed.date()
    return None


def _looks_like_date_label(val) -> bool:
    if isinstance(val, (datetime, date)):
        return True
    if not isinstance(val, str):
        return False
    s = val.strip()
    return bool(re.match(r"^\d{1,4}[./-]\d{1,2}(?:[./-]\d{1,4})?$", s))


def _coerce_date_series(values) -> pd.Series:
    index = values.index if hasattr(values, "index") else None
    return pd.to_datetime(
        pd.Series(values, index=index).map(lambda v: _parse_sheet_date(v) or v),
        errors="coerce",
        dayfirst=True,
    )


VISUAL_SHEETS = ["APR26", "MAR26", "JS APR26", "JS-FOR MAR26", "FOR Trades"]

# Sheets matching this pattern are treated as month-trade sheets and exposed
# as filterable months in the dashboard. Examples: APR26, MAR26, JS APR26.
# Future months (MAY26, JUN26, JS MAY26, ...) are picked up automatically.
MONTH_SHEET_REGEX = re.compile(r"^(JS )?[A-Z]{3}\d{2}$")

# Optional metadata tab listing forward-test strategies by id (auto-derived
# from each trade's name + underlying). Forward trades themselves live in the
# regular `FOR Trades` tab — this sheet only enriches display.
FT_STRATEGIES_SHEET = "FT Strategies"

TRADE_UNDERLYINGS = (
    "SPX",
    "SPXW",
    "XSP",
    "SPY",
    "NDX",
    "NDXP",
    "QQQ",
    "RUT",
    "IWM",
    "DIA",
    "GLD",
    "SLV",
    "XLB",
    "XLC",
    "XLE",
    "XLF",
    "XLI",
    "XLK",
    "XLP",
    "XLRE",
    "XLU",
    "XLV",
    "XLY",
    "BTC",
    "BITCOIN",
)

UNDERLYING_ALIASES = {
    "SPXW": "SPX",
    "NDXP": "NDX",
    "BITCOIN": "BTC",
}


def infer_visual_sheet_env(sheet_name: str) -> str:
    return SHEET_ENV_MAP.get(
        sheet_name,
        "JS_Forward" if sheet_name.startswith("JS ") else "CZ_Live",
    )


def iter_visual_sheet_names(wb) -> list[str]:
    return [
        sheet_name
        for sheet_name in wb.sheetnames
        if MONTH_SHEET_REGEX.match(sheet_name) or sheet_name in VISUAL_SHEETS
    ]


# ─── Google Drive download ────────────────────────────────────────────────────

def _write_json_env_to_file(env_name: str, path: Path) -> None:
    raw = os.environ.get(env_name, "").strip()
    if not raw:
        return
    parsed = json.loads(raw)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(parsed), encoding="utf-8")


def materialize_gdrive_env_files() -> None:
    """Allows cloud deployments to provide OAuth files as secret env vars."""
    _write_json_env_to_file("GDRIVE_CREDENTIALS_JSON", GDRIVE_CREDS)
    _write_json_env_to_file("GDRIVE_TOKEN_JSON", GDRIVE_TOKEN)


def can_run_browser_oauth() -> bool:
    if os.getenv("GDRIVE_ALLOW_BROWSER_AUTH", "").strip().lower() in ("1", "true", "yes", "on"):
        return True
    if any(os.getenv(name) for name in NON_INTERACTIVE_ENV_VARS):
        return False
    return sys.stdin.isatty()


def _quote_sheet_title(title: str) -> str:
    return "'" + title.replace("'", "''") + "'"


def _sheet_values_range(title: str) -> str:
    return f"{_quote_sheet_title(title)}!A:ZZ"


def _write_values_workbook(value_ranges: list[dict], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    for value_range in value_ranges:
        range_name = str(value_range.get("range") or "")
        title = range_name.split("!", 1)[0].strip("'").replace("''", "'")
        ws = wb.create_sheet(title[:31])
        for row_values in value_range.get("values") or []:
            ws.append(row_values)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def download_sheet_values(file_id: str, output_path: Path, creds) -> bool:
    """Fallback for Google Sheets that exceed the Drive XLSX export limit."""
    from googleapiclient.discovery import build

    sheets = build("sheets", "v4", credentials=creds)
    meta = sheets.spreadsheets().get(
        spreadsheetId=file_id,
        fields="sheets(properties(title))",
        includeGridData=False,
    ).execute()
    titles = [
        item["properties"]["title"]
        for item in meta.get("sheets", [])
        if item.get("properties", {}).get("title")
    ]
    needed = [
        title
        for title in titles
        if title in {"db_robots", "db_cria"} or MONTH_SHEET_REGEX.match(title) or title in VISUAL_SHEETS
    ]
    if not needed:
        print("[!] Sheets API fallback nao encontrou abas necessarias.")
        return False

    print(f"[export] Sheets API fallback: lendo {len(needed)} abas ({', '.join(needed)})")
    value_ranges = sheets.spreadsheets().values().batchGet(
        spreadsheetId=file_id,
        ranges=[_sheet_values_range(title) for title in needed],
        valueRenderOption="UNFORMATTED_VALUE",
        dateTimeRenderOption="FORMATTED_STRING",
        majorDimension="ROWS",
    ).execute().get("valueRanges", [])
    _write_values_workbook(value_ranges, output_path)
    return True


def download_from_gdrive(file_id: str, output_path: Path) -> bool:
    """
    Baixa o Google Sheet como XLSX via Drive API (OAuth2).
    Primeira execução: abre browser para consentimento.
    Execuções seguintes: usa token cached (auto-refresh).
    """
    try:
        materialize_gdrive_env_files()
    except Exception as e:
        print(f"[!] Falha ao materializar credenciais Google Drive via env: {e}")

    if not GDRIVE_CREDS.exists():
        print(f"[!] Credenciais nao encontradas em: {GDRIVE_CREDS}")
        print("    Baixe o OAuth JSON do GCP e salve nesse caminho.")
        return False
    try:
        from google.auth.transport.requests import Request
        from google.auth.exceptions import RefreshError
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from googleapiclient.discovery import build
        from googleapiclient.errors import HttpError

        creds = None
        if GDRIVE_TOKEN.exists():
            creds = Credentials.from_authorized_user_file(str(GDRIVE_TOKEN), GDRIVE_SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                try:
                    creds.refresh(Request())
                except RefreshError as e:
                    if not can_run_browser_oauth():
                        print(f"[!] Google Drive token expirado/revogado em ambiente non-interactive: {e}")
                        print("    Recrie GDRIVE_TOKEN_JSON e atualize o secret no Render.")
                        return False
                    print(f"[!] Google Drive token expirado/revogado; abrindo OAuth local: {e}")
                    creds = None
            else:
                creds = None

            if not creds:
                if not can_run_browser_oauth():
                    print("[!] Google Drive token ausente/invalido em ambiente non-interactive.")
                    print("    Configure GDRIVE_CREDENTIALS_JSON e GDRIVE_TOKEN_JSON no Render.")
                    return False
                flow = InstalledAppFlow.from_client_secrets_file(str(GDRIVE_CREDS), GDRIVE_SCOPES)
                creds = flow.run_local_server(port=0)

            GDRIVE_TOKEN.parent.mkdir(parents=True, exist_ok=True)
            GDRIVE_TOKEN.write_text(creds.to_json())

        service = build("drive", "v3", credentials=creds)
        try:
            content = service.files().export_media(
                fileId=file_id,
                mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ).execute()
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_bytes(content)
        except HttpError as e:
            reason = ""
            try:
                reason = (e.error_details or [{}])[0].get("reason", "")
            except Exception:
                reason = ""
            if reason != "exportSizeLimitExceeded":
                raise
            print("[!] Drive export bloqueado por tamanho; usando Sheets API fallback.")
            return download_sheet_values(file_id, output_path, creds)
        return True
    except Exception as e:
        print(f"[!] Google Drive download falhou: {e}")
        return False


# ─── Sheet summaries (lê TOTAL rows das abas visuais) ────────────────────────

def read_sheet_summaries(xlsx_path: Path) -> dict[str, list[dict]]:
    """
    Lê a linha TOTAL de cada aba visual — fonte autoritativa de Open PnL e RLZD.

    Estrutura de colunas (0-indexed, consistente em todas as abas):
      col 1 (B) = label  ← procurar "TOTAL"
      col 4 (E) = MxProf
      col 5 (F) = Open PnL
      col 7 (H) = RLZD
      col 8 (I) = Delta  (nem todas as abas têm)

    Retorna: {env_norm: [{month, open_pnl, rlzd, delta, max_profit}]}
    """
    result: dict[str, list[dict]] = {}
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    for sheet_name in iter_visual_sheet_names(wb):
        env_norm = infer_visual_sheet_env(sheet_name)
        ws = wb[sheet_name]

        total_row = None
        for row in ws.iter_rows(min_row=1, max_row=25, values_only=True):
            if len(row) > 1 and row[1] is not None:
                if str(row[1]).strip().upper() == "TOTAL":
                    total_row = row
                    break

        if total_row is None:
            continue

        def _s(idx):
            return total_row[idx] if len(total_row) > idx else None

        open_pnl   = _parse_money(_s(5))
        rlzd       = _parse_money(_s(7))
        max_profit = _parse_money(_s(4))
        delta_raw  = _s(8)
        delta      = _parse_money(delta_raw) if isinstance(delta_raw, (int, float)) else None

        if env_norm not in result:
            result[env_norm] = []
        result[env_norm].append({
            "month":      sheet_name,
            "open_pnl":   open_pnl  if open_pnl  is not None else 0.0,
            "rlzd":       rlzd      if rlzd       is not None else 0.0,
            "delta":      delta     if delta      is not None else 0.0,
            "max_profit": max_profit if max_profit is not None else 0.0,
        })

    wb.close()
    return result


def read_individual_trade_pnls(xlsx_path: Path) -> tuple[dict[str, float], dict[str, list[str]], dict[str, dict], dict[str, list[dict]]]:
    """
    For every sheet matching MONTH_SHEET_REGEX (APR26, MAR26, JS APR26, future
    MAY26 etc.), reads each trade's last PnL directly from the visual block AND
    records which trades live in which sheet.

    Block layout (0-indexed within the row tuple):
      name_row (row 4) : trade name at block_start col
      pnl_row          : first row where bc+1=DIT, bc+2=DTE, bc+3=PnL are ALL numeric

    Returns (
        {normalize_name(trade_name): pnl},
        {sheet_name: [trade_names_in_that_sheet]},
        {normalize_name(trade_name): visual_metadata},
    )
    """
    result: dict[str, float] = {}
    sheet_to_trades: dict[str, list[str]] = {}
    trade_visual_details: dict[str, dict] = {}
    sheet_daily_pnls: dict[str, list[dict]] = {}

    def is_num(v):
        return isinstance(v, (int, float)) and not isinstance(v, bool)

    def as_date(v) -> date | None:
        return _parse_sheet_date(v)

    def as_date_str(v) -> str | None:
        d = as_date(v)
        return d.isoformat() if d else None

    def weekday_str(d: date | None) -> str | None:
        return d.strftime("%a") if d else None

    def calendar_days(start: date | None, end: date | None) -> int | None:
        if not start or not end or end < start:
            return None
        return (end - start).days

    def date_header_location(rows: dict[int, tuple]) -> tuple[int | None, int | None]:
        for r, row_data in rows.items():
            for ci, val in enumerate(row_data):
                if val is not None and str(val).strip().lower() == "date":
                    return r, ci
        return None, None

    def sheet_last_update(rows: dict[int, tuple]) -> date | None:
        for row_data in rows.values():
            for ci, val in enumerate(row_data):
                if val is not None and str(val).strip().lower().startswith("last update"):
                    for maybe_date in row_data[ci + 1 : ci + 6]:
                        parsed = as_date(maybe_date)
                        if parsed:
                            return parsed
        return None

    def last_pnl_mark(ws, start_row: int | None, date_col: int | None, pnl_col: int) -> dict | None:
        if start_row is None or date_col is None:
            return None
        last = None
        for r_idx, row_data in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
            row_date = as_date(row_data[date_col] if len(row_data) > date_col else None)
            pnl_val = row_data[pnl_col] if len(row_data) > pnl_col else None
            if row_date and is_num(pnl_val):
                last = {
                    "row": r_idx,
                    "date": row_date,
                    "pnl": float(pnl_val),
                }
        if not last:
            return None
        return {
            "row": last["row"],
            "date_obj": last["date"],
            "date": last["date"].isoformat(),
            "weekday": weekday_str(last["date"]),
            "pnl": last["pnl"],
        }

    def daily_sheet_pnl(
        ws,
        start_row: int | None,
        date_col: int | None,
        blocks: list[tuple[int, str]],
        cutoff: date | None,
        status_row: tuple,
        info_row: tuple,
    ) -> list[dict]:
        if start_row is None or date_col is None:
            return []
        last_marks = {
            pnl_col: last_pnl_mark(ws, start_row, date_col, pnl_col)
            for pnl_col, _ in blocks
        }

        def is_finalized(pnl_col: int, row_date: date, final_date: date | None) -> bool:
            status_val = status_row[pnl_col] if len(status_row) > pnl_col else None
            status = str(status_val).strip().lower() if status_val is not None else ""
            exp_dt = as_date(info_row[pnl_col + 2] if len(info_row) > pnl_col + 2 else None)
            status_closed = "closed" in status
            expired = exp_dt is not None and row_date >= exp_dt
            has_final_mark = final_date is not None and row_date >= final_date
            return has_final_mark and (status_closed or expired)

        rows: list[dict] = []
        prev_total: float | None = None
        for r_idx, row_data in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
            row_date = as_date(row_data[date_col] if len(row_data) > date_col else None)
            if not row_date:
                continue
            if cutoff and row_date > cutoff:
                continue
            trade_pnls = []
            open_total = 0.0
            rlzd_total = 0.0
            for pnl_col, trade_name in blocks:
                pnl_val = row_data[pnl_col] if len(row_data) > pnl_col else None
                final_mark = last_marks.get(pnl_col)
                final_date = final_mark.get("date_obj") if final_mark else None

                if not is_num(pnl_val):
                    if not (final_mark and final_date and is_finalized(pnl_col, row_date, final_date)):
                        continue
                    pnl = float(final_mark["pnl"])
                    bucket = "rlzd"
                    rlzd_total += pnl
                else:
                    pnl = float(pnl_val)
                    if is_finalized(pnl_col, row_date, final_date):
                        bucket = "rlzd"
                        rlzd_total += pnl
                    else:
                        bucket = "open"
                        open_total += pnl

                trade_pnls.append({"name": trade_name, "pnl": round(pnl, 2), "bucket": bucket})

            if trade_pnls:
                total = round(open_total + rlzd_total, 2)
                daily_change = total if prev_total is None else total - prev_total
                prev_total = total
                rows.append({
                    "date": row_date.isoformat(),
                    "row": r_idx,
                    "pnl": total,
                    "open_pnl": round(open_total, 2),
                    "rlzd": round(rlzd_total, 2),
                    "daily_pnl": round(daily_change, 2),
                    "pnl_type": "total_open_plus_rlzd",
                    "n_trades": len(trade_pnls),
                    "trades": trade_pnls,
                })
        return rows

    def last_dte_bucket(row: tuple, block_col: int) -> str | None:
        bucket = None
        for val in row[: block_col + 1]:
            if val is None:
                continue
            s = str(val).strip()
            if "DTE" in s.upper():
                bucket = s
        return bucket

    SKIP_LABELS = {
        "active",
        "closed",
        "live",
        "name",
        "new",
        "open",
        "options control panel",
        "status",
        "strategy",
        "titel",
        "trade",
        "url",
    }
    VALID_STATUSES = {"active", "closed", "live", "new", "open"}

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    block_sheets = [s for s in wb.sheetnames if MONTH_SHEET_REGEX.match(s)]
    print(f"  [trade-sheets detected] {block_sheets}")

    for sheet_name in block_sheets:
        ws = wb[sheet_name]

        # Read first 25 rows into dict keyed by 1-based row number
        all_rows: dict[int, tuple] = {}
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True), start=1):
            all_rows[r_idx] = row
        date_header_r, date_col = date_header_location(all_rows)
        pnl_series_start_r = date_header_r + 1 if date_header_r is not None else None
        last_update_dt = sheet_last_update(all_rows)

        # Locate the row containing "Titel" — this is the trade-name row
        name_r = None
        for r, row_data in all_rows.items():
            if any(v is not None and str(v).strip().lower() == "titel" for v in row_data):
                name_r = r
                break

        title_layout = name_r is not None
        if name_r is None:
            # Legacy visual sheets place trade names on row 3 and statuses on row 2.
            best_r = None
            best_score = -1
            for candidate_r in range(2, 7):
                row_data = all_rows.get(candidate_r, ())
                status_row = all_rows.get(candidate_r - 1, ())
                score = 0
                for ci, val in enumerate(row_data):
                    if val is None or _looks_like_date_label(val):
                        continue
                    s = str(val).strip()
                    if not s or s.lower() in SKIP_LABELS or s.startswith("http"):
                        continue
                    try:
                        float(s)
                        continue
                    except ValueError:
                        pass
                    st_val = status_row[ci] if len(status_row) > ci else None
                    st_str = str(st_val).strip().lower() if st_val is not None else ""
                    if st_str and st_str not in VALID_STATUSES:
                        continue
                    score += 1
                if score > best_score:
                    best_r = candidate_r
                    best_score = score
            name_r = best_r

        name_row = all_rows[name_r]

        # Status row is 2 rows above Titel row (row 2 when name_r = 4)
        status_r   = name_r - (2 if title_layout else 1)
        status_row = all_rows.get(status_r, ())
        VALID_STATUSES = {"active", "closed", "live", "new", "open"}

        # Collect block start columns: cols with trade names in name_row
        # Guard: same column in the status row must have a known status value —
        # this eliminates labels like "Last Update:" and date cells.
        blocks: list[tuple[int, str]] = []
        for ci, val in enumerate(name_row):
            if val is None:
                continue
            # Skip date labels returned by openpyxl or the Sheets API fallback.
            if _looks_like_date_label(val):
                continue
            s = str(val).strip()
            if not s or s.lower() in SKIP_LABELS or s.startswith("http"):
                continue
            try:
                float(s)
                continue  # pure number → not a trade name
            except ValueError:
                pass
            # Cross-check: status cell at same column must be blank (= active)
            # or a known status keyword. Anything else (e.g. "OPTIONS CONTROL PANEL")
            # means this column is not a trade block.
            st_val = status_row[ci] if len(status_row) > ci else None
            st_str = str(st_val).strip().lower() if st_val is not None else ""
            if st_str and st_str not in VALID_STATUSES:
                continue
            blocks.append((ci, normalize_name(s)))

        if not blocks:
            print(f"  [{sheet_name}] no trade blocks found")
            continue

        info_r = 6 if title_layout else 5
        strikes_r = 7 if title_layout else 6
        info_row = all_rows.get(info_r, ())
        strikes_row = all_rows.get(strikes_r, ())

        # Record every detected trade name for this sheet (membership map)
        sheet_to_trades.setdefault(sheet_name, []).extend(name for _, name in blocks)
        sheet_daily_pnls[sheet_name] = daily_sheet_pnl(
            ws,
            pnl_series_start_r,
            date_col,
            blocks,
            last_update_dt,
            status_row,
            info_row,
        )

        # For each block, find the DIT|DTE|PnL row (all three offsets numeric)
        for bc, trade_name in blocks:
            status_val = status_row[bc] if len(status_row) > bc else None
            status = str(status_val).strip() if status_val is not None and str(status_val).strip() else "Active"
            open_price = info_row[bc] if len(info_row) > bc else None
            open_dt = info_row[bc + 1] if len(info_row) > bc + 1 else None
            exp_dt = info_row[bc + 2] if len(info_row) > bc + 2 else None
            dte_open = info_row[bc + 3] if len(info_row) > bc + 3 else None
            strikes = strikes_row[bc] if len(strikes_row) > bc else None
            open_date_obj = as_date(open_dt)
            last_mark = last_pnl_mark(ws, pnl_series_start_r, date_col, bc)
            last_mark_date = last_mark.get("date_obj") if last_mark else None
            status_is_closed = "closed" in status.lower()
            days_open_as_of_last_pnl = calendar_days(open_date_obj, last_mark_date)
            days_held = days_open_as_of_last_pnl if status_is_closed else None
            pnl = None
            if title_layout:
                for r in range(name_r + 2, min(name_r + 22, 26)):
                    row_data = all_rows.get(r, ())
                    dit = row_data[bc + 1] if len(row_data) > bc + 1 else None
                    dte = row_data[bc + 2] if len(row_data) > bc + 2 else None
                    p   = row_data[bc + 3] if len(row_data) > bc + 3 else None
                    if is_num(dit) and is_num(dte) and is_num(p):
                        pnl = float(p)
                        break
            else:
                # Legacy layout: row 16 labels the per-trade time-series columns.
                # The final trade PnL is the last numeric value in the block's PnL column.
                if last_mark:
                    pnl = float(last_mark["pnl"])
            if pnl is None and last_mark:
                pnl = float(last_mark["pnl"])

            if pnl is not None:
                result[trade_name] = pnl
                trade_visual_details[trade_name] = {
                    "sheet": sheet_name,
                    "status": status,
                    "dte_bucket": last_dte_bucket(all_rows.get(1, ()), bc),
                    "open_price": float(open_price) if is_num(open_price) else None,
                    "open_date": as_date_str(open_dt),
                    "exp_date": as_date_str(exp_dt),
                    "dte_open": int(dte_open) if is_num(dte_open) else None,
                    "strikes": str(strikes).strip() if strikes is not None and str(strikes).strip() else None,
                    "pnl": pnl,
                    "last_pnl": float(last_mark["pnl"]) if last_mark else None,
                    "last_pnl_date": last_mark.get("date") if last_mark else None,
                    "last_pnl_weekday": last_mark.get("weekday") if last_mark else None,
                    "last_pnl_row": last_mark.get("row") if last_mark else None,
                    "inferred_close_date": last_mark.get("date") if status_is_closed and last_mark else None,
                    "inferred_close_weekday": last_mark.get("weekday") if status_is_closed and last_mark else None,
                    "days_held": days_held,
                    "days_open_as_of_last_pnl": days_open_as_of_last_pnl,
                }
                print(f"  [{sheet_name}] {trade_name:22s}: ${pnl:+,.2f}")
            else:
                print(f"  [{sheet_name}] {trade_name:22s}: PnL row not found — db_robots ffill used")

    wb.close()
    return result, sheet_to_trades, trade_visual_details, sheet_daily_pnls


def load_closed_trades_from_sheets(xlsx_path: Path) -> set[str]:
    """
    Lê as abas visuais e retorna o set de nomes de trades marcados como Closed
    (row 2 do bloco = 'Closed'). O nome vem da row 4 da mesma coluna.
    """
    closed = set()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    for sheet_name in iter_visual_sheet_names(wb):
        ws = wb[sheet_name]
        # Ler rows 2-6 de uma vez (templates diferem entre abas:
        # APR26 tem nome em row4, MAR26 tem nome em row3)
        rows = {r: list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
                for r in range(2, 7)}

        n_cols = max(len(rows[r]) for r in rows)
        for col_idx in range(n_cols):
            status = rows[2][col_idx] if col_idx < len(rows[2]) else None
            if not status or "CLOSED" not in str(status).upper():
                continue

            # Procurar nome do trade nas rows 3, 4, 5 (primeira que não é URL nem número)
            name = None
            for r in [3, 4, 5]:
                v = rows[r][col_idx] if col_idx < len(rows[r]) else None
                if not v:
                    continue
                s = str(v).strip()
                if s and not s.startswith("http") and not s.startswith("www"):
                    try:
                        float(s)   # se for número puro, ignorar
                    except ValueError:
                        name = s
                        break

            if name:
                closed.add(normalize_name(name))

    wb.close()
    return closed


def load_db_robots(xlsx_path: Path) -> pd.DataFrame:
    """Carrega db_robots ignorando as linhas de placeholder (env vazio)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb["db_robots"]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        env = row[1] if len(row) > 1 else None
        if not env or not str(env).strip():
            continue
        strategy_raw = row[2] if len(row) > 2 else None
        if not strategy_raw or not str(strategy_raw).strip():
            continue

        raw = str(strategy_raw).strip()
        # Detecta se o trade foi marcado como fechado no Make
        # (o bot escreve " - CLOSED" ou " - CLOS" no nome quando fecha)
        was_closed = " - CLOSED" in raw.upper() or " - CLOS" in raw.upper()

        rows.append({
            "date":           row[0],
            "environment":    str(env).strip(),
            "strategy":       normalize_name(raw),
            "strategy_raw":   raw,
            "was_closed_flag": was_closed,
            "pnl":            _parse_money(row[3]),
            "delta":          _parse_money(row[4]),
        })

    wb.close()
    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df["date"] = _coerce_date_series(df["date"])
    df.dropna(subset=["date"], inplace=True)
    df["env_norm"] = df["environment"].map(lambda e: normalize_env(e))
    df.sort_values("date", inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df


def load_db_cria(xlsx_path: Path) -> pd.DataFrame:
    """Carrega db_cria (dados de abertura de cada trade)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb["db_cria"]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        def _get(idx):
            return row[idx] if len(row) > idx else None

        dte_raw, dte_int = _parse_dte(_get(2))
        rows.append({
            "scraped_date":     _get(0),
            "url":              str(_get(1)) if _get(1) else None,
            "dte_open":         dte_int,
            "dte_open_raw":     dte_raw,
            "sd":               _parse_money(_get(3)),
            "open_date":        _get(4),
            "underlying_price": _parse_money(_get(5)),
            "soll_lw_be":       _parse_money(_get(6)),
            "soll_up_be":       _parse_money(_get(7)),
            "net_credit":       _parse_money(_get(8)),
            "max_loss":         _parse_money(_get(9)),
            "strikes":          str(_get(10)) if _get(10) else None,
            "ist_lw_be":        _parse_money(_get(11)),
            "ist_up_be":        _parse_money(_get(12)),
            "trade_name":       normalize_name(str(_get(13))) if _get(13) else None,
        })

    wb.close()
    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df["scraped_date"] = _coerce_date_series(df["scraped_date"])
    df["open_date"]    = _coerce_date_series(df["open_date"])
    df.dropna(subset=["trade_name"], inplace=True)
    df.sort_values("scraped_date", inplace=True)
    # Keep last entry per trade_name (most recent scrape wins)
    df = df.drop_duplicates(subset=["trade_name"], keep="last")
    df.reset_index(drop=True, inplace=True)
    return df


# ─── ForwardTest readers ─────────────────────────────────────────────────────

def load_ft_strategies(xlsx_path: Path) -> list[dict]:
    """Optional metadata tab — enriches forward-test cards with descriptions, leg
    templates, and per-strategy start_date overrides. Strategy assignment itself
    is handled at API time by parsing each trade name (see `strategy_family`).

    Schema (all columns optional except A and B):
      A=strategy_id (e.g., `triple-calendar_iwm`),
      B=name (display label),
      C=description, D=horizon, E=legs_template (JSON),
      F=entry_rule, G=exit_rule, H=start_date (YYYY-MM-DD), I=status.
    Returns [] when the tab is missing.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if FT_STRATEGIES_SHEET not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[FT_STRATEGIES_SHEET]

    strategies: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        strategy_id = row[0] if len(row) > 0 else None
        name = row[1] if len(row) > 1 else None
        if not strategy_id or not str(strategy_id).strip():
            continue
        if not name or not str(name).strip():
            continue

        legs_raw = row[4] if len(row) > 4 else None
        legs_template: list[dict] = []
        if legs_raw and str(legs_raw).strip():
            try:
                parsed = json.loads(str(legs_raw))
                if isinstance(parsed, list):
                    legs_template = parsed
            except json.JSONDecodeError as exc:
                print(f"  [FT Strategies] {strategy_id}: legs_template JSON inválido — {exc}")

        start_date_iso: str | None = None
        if len(row) > 7 and row[7]:
            d = _parse_sheet_date(row[7])
            if d:
                start_date_iso = d.isoformat()

        strategies.append({
            "strategy_id": str(strategy_id).strip(),
            "name": str(name).strip(),
            "description": str(row[2]).strip() if len(row) > 2 and row[2] else None,
            "horizon": str(row[3]).strip() if len(row) > 3 and row[3] else None,
            "legs_template": legs_template,
            "entry_rule": str(row[5]).strip() if len(row) > 5 and row[5] else None,
            "exit_rule": str(row[6]).strip() if len(row) > 6 and row[6] else None,
            "start_date": start_date_iso,
            "status": str(row[8]).strip().lower() if len(row) > 8 and row[8] else "active",
        })

    wb.close()
    strategies.sort(key=lambda s: s["name"])
    return strategies


_FORWARD_ENVS = ("CZ_Forward", "JS_Forward")


def _attach_forward_daily_history(trades: list[dict], db_robots: pd.DataFrame) -> None:
    """Embed per-trade daily PnL/Delta series for forward trades.

    Forward trades (CZ_Forward and JS_Forward) drive the per-trade Journey/Delta
    charts and milestone calculations in the app. History is sourced from
    db_robots — same canonical source that powers the visual sheets via
    formulas. Mutates `trades` in place.
    """
    if db_robots.empty:
        return
    forward_names = {t["name"] for t in trades if t.get("environment") in _FORWARD_ENVS}
    if not forward_names:
        return

    rows = db_robots[db_robots["strategy"].isin(forward_names)].sort_values("date")
    history_by_name: dict[str, list[dict]] = {}
    for name, grp in rows.groupby("strategy"):
        series: list[dict] = []
        for _, r in grp.iterrows():
            d = r.get("date")
            date_iso = d.strftime("%Y-%m-%d") if pd.notna(d) else None
            pnl = r.get("pnl")
            delta = r.get("delta")
            series.append({
                "date": date_iso,
                "pnl": float(pnl) if pd.notna(pnl) else None,
                "delta": float(delta) if pd.notna(delta) else None,
            })
        history_by_name[str(name)] = series

    for t in trades:
        if t.get("environment") in _FORWARD_ENVS:
            t["daily_history"] = history_by_name.get(t["name"], [])


# ─── Derivar estado atual dos trades ─────────────────────────────────────────

def build_trade_snapshot(db_robots: pd.DataFrame, db_cria: pd.DataFrame,
                         closed_from_sheets: set[str] | None = None) -> list[dict]:
    """
    Combina db_robots (histórico PnL/Delta) e db_cria (dados de abertura)
    para produzir o estado atual de cada trade.
    """
    if db_robots.empty:
        return []

    today      = pd.Timestamp.today().normalize()
    cutoff     = today - pd.Timedelta(days=4)   # activo se teve update nos últimos 4 dias

    # Último registro por trade (nome)
    latest = (
        db_robots.sort_values("date")
        .groupby("strategy", as_index=False)
        .last()
    )

    # Série de datas por trade para calcular RLZD aproximado
    # (primeiro PnL = referência; ultimo PnL = atual)
    first_pnl = (
        db_robots.dropna(subset=["pnl"])
        .groupby("strategy")["pnl"]
        .first()
        .rename("pnl_first")
    )

    # Trades marcados como CLOSED: três fontes em ordem de confiança:
    # 1. Abas visuais (row 2 = "Closed") — fonte mais confiável
    # 2. db_robots com "- CLOSED" no nome — marcação pelo Make
    closed_from_sheets = closed_from_sheets or set()
    closed_by_flag = set(
        db_robots[db_robots["was_closed_flag"]]["strategy"].unique()
    )
    all_closed = closed_from_sheets | closed_by_flag

    trades = []
    for _, row in latest.iterrows():
        name     = row["strategy"]
        env_raw  = row["environment"]
        env_norm = row["env_norm"]
        last_dt  = row["date"]

        # Trade names prefixed with "JS " override env to JS_Forward, regardless
        # of which Environment column was used in db_robots. Lets João log JS
        # forward trades using the "FOR Trades" env tag without losing the user
        # split downstream.
        if isinstance(name, str) and name.upper().startswith("JS ") and env_norm in ("CZ_Forward", "CZ_Live"):
            env_norm = "JS_Forward"

        manually_closed = name in all_closed
        is_active = (not manually_closed) and (last_dt >= cutoff)
        # Ainda ativo pela data? Checar se já expirou pelo DTE calculado
        # (será corrigido após calcular exp_date abaixo — reavaliado no final)

        # Dados de abertura do db_cria
        cria_rows = db_cria[db_cria["trade_name"] == name]
        cria = cria_rows.iloc[0].to_dict() if not cria_rows.empty else {}

        net_credit = cria.get("net_credit")
        max_loss   = cria.get("max_loss")
        dte_open       = cria.get("dte_open")
        dte_open_raw   = cria.get("dte_open_raw")
        # Fallback: derive DTE from trade name when db_cria didn't capture it
        # (e.g. "FOR01 RUT TripleCalendar 14/21DTE", "T03 SLV Triple Calendar 11/18DTE").
        if not dte_open_raw and not dte_open and isinstance(name, str):
            m = re.search(r"(\d+(?:\s*[/-]\s*\d+)?)\s*DTE", name, re.IGNORECASE)
            if m:
                dte_open_raw = re.sub(r"\s+", "", m.group(1))
                first = re.search(r"\d+", dte_open_raw)
                if first:
                    dte_open = int(first.group())
        open_date  = cria.get("open_date")
        und_price  = cria.get("underlying_price")
        lw_be      = cria.get("ist_lw_be") or cria.get("soll_lw_be")
        up_be      = cria.get("ist_up_be") or cria.get("soll_up_be")

        # Calcular DTE restante
        exp_date = None
        dte_remaining = None
        if open_date and dte_open:
            try:
                exp_date = open_date + pd.Timedelta(days=int(dte_open))
                dte_remaining = max(0, (exp_date - today).days)
                # Trade com DTE calculado = 0 já expirou — não é ativo
                if dte_remaining == 0 and is_active:
                    is_active = False
            except Exception:
                pass

        # PnL%
        pnl_current = row.get("pnl")
        pnl_pct_max = None
        if pnl_current is not None and net_credit and net_credit != 0:
            pnl_pct_max = round(pnl_current / net_credit * 100, 1)

        # % spot para BEs
        pct_to_lw = None
        pct_to_up = None
        if und_price and und_price > 0:
            if lw_be:
                pct_to_lw = round((lw_be - und_price) / und_price * 100, 2)
            if up_be:
                pct_to_up = round((up_be - und_price) / und_price * 100, 2)

        # Inferir underlying (SPX/NDX/RUT/etc.) pelo nome
        underlying = _infer_underlying(name)

        trades.append({
            "name":              name,
            "environment_raw":   env_raw,
            "environment":       env_norm,
            "underlying":        underlying,
            "is_active":         is_active,
            "last_update":       last_dt.strftime("%Y-%m-%d") if pd.notna(last_dt) else None,
            # Abertura
            "open_date":         open_date.strftime("%Y-%m-%d") if pd.notna(open_date) else None,
            "exp_date":          exp_date.strftime("%Y-%m-%d") if exp_date and pd.notna(exp_date) else None,
            "dte_open":          int(dte_open) if pd.notna(dte_open) and dte_open else None,
            "dte_open_raw":      dte_open_raw if isinstance(dte_open_raw, str) and dte_open_raw else None,
            "dte_remaining":     dte_remaining,
            "underlying_price_at_open": und_price,
            "strikes":           cria.get("strikes"),
            "net_credit":        net_credit,
            "max_loss":          max_loss,
            "sd":                cria.get("sd"),
            "lw_be":             lw_be,
            "up_be":             up_be,
            "pct_to_lw_be":      pct_to_lw,
            "pct_to_up_be":      pct_to_up,
            # Estado atual
            "pnl_current":       pnl_current,
            "pnl_pct_max":       pnl_pct_max,
            "delta_current":     row.get("delta"),
            # Alertas
            "alert_dte":         (dte_remaining is not None and dte_remaining <= 7),
            "alert_profit_50":   (pnl_pct_max is not None and pnl_pct_max >= 50),
            "alert_stop":        (pnl_pct_max is not None and pnl_pct_max <= -100),
            "tent_status":       _tent_status(pnl_pct_max),
        })

    return sorted(trades, key=lambda t: (not t["is_active"], t["name"]))


def _infer_underlying(name: str) -> str:
    n = name.upper()
    for sym in TRADE_UNDERLYINGS:
        if re.search(rf"(?<![A-Z0-9]){re.escape(sym)}(?![A-Z0-9])", n):
            return UNDERLYING_ALIASES.get(sym, sym)
    return "?"


def _infer_underlying_from_visual(name: str, open_price: float | None) -> str:
    underlying = _infer_underlying(name)
    if underlying != "?":
        return underlying
    if open_price is None:
        return underlying
    if open_price > 15000:
        return "NDX"
    if open_price > 4000:
        return "SPX"
    if open_price > 1000:
        return "RUT"
    return underlying


def _tent_status(pnl_pct: float | None) -> str:
    if pnl_pct is None:
        return "unknown"
    if pnl_pct >= 50:
        return "profit_target"
    if pnl_pct <= -100:
        return "stop_loss"
    if pnl_pct < 0:
        return "loss"
    return "profit"


# ─── Monthly Summary ──────────────────────────────────────────────────────────

def build_monthly_summary(db_robots: pd.DataFrame, db_cria: pd.DataFrame) -> pd.DataFrame:
    """Resume performance por mês e por trade."""
    if db_robots.empty:
        return pd.DataFrame()

    # PnL first e last por trade (aproximação de RLZD)
    grp = db_robots.dropna(subset=["pnl"]).groupby("strategy")
    summary = pd.DataFrame({
        "first_date":    grp["date"].first(),
        "last_date":     grp["date"].last(),
        "pnl_first":     grp["pnl"].first(),
        "pnl_last":      grp["pnl"].last(),
        "environment":   grp["env_norm"].last(),
    }).reset_index()

    summary["month"] = summary["first_date"].dt.strftime("%b%y").str.upper()
    summary["dit"]   = (summary["last_date"] - summary["first_date"]).dt.days

    # Juntar db_cria
    if not db_cria.empty:
        cria_sub = db_cria[["trade_name", "net_credit", "max_loss", "dte_open", "underlying_price"]].copy()
        summary = summary.merge(cria_sub, left_on="strategy", right_on="trade_name", how="left")

    summary["pnl_pct_max"] = np.where(
        summary["net_credit"].notna() & (summary["net_credit"] != 0),
        (summary["pnl_last"] / summary["net_credit"] * 100).round(1),
        None
    )
    summary["won"] = (summary["pnl_last"] > 0).astype(int)

    return summary.sort_values(["month", "strategy"]).reset_index(drop=True)


# ─── Trade history parquet ────────────────────────────────────────────────────

def build_trade_history(db_robots: pd.DataFrame, db_cria: pd.DataFrame) -> pd.DataFrame:
    """Retorna série temporal completa de PnL/Delta por trade."""
    df = db_robots.copy()
    if df.empty:
        return df

    # Enriquecer com dados de abertura
    if not db_cria.empty:
        cria_sub = db_cria[["trade_name", "net_credit", "max_loss", "open_date", "dte_open"]].copy()
        df = df.merge(cria_sub, left_on="strategy", right_on="trade_name", how="left")

    df["pnl_pct_max"] = np.where(
        df["net_credit"].notna() & (df["net_credit"] != 0),
        (df["pnl"] / df["net_credit"] * 100).round(2),
        None
    )
    return df


# ─── Portfolio KPIs ───────────────────────────────────────────────────────────

def compute_portfolio_kpis(trades: list[dict], monthly: pd.DataFrame) -> dict:
    active   = [t for t in trades if t["is_active"]]
    inactive = [t for t in trades if not t["is_active"]]

    def safe_sum(lst, key):
        vals = [t[key] for t in lst if t.get(key) is not None]
        return round(sum(vals), 2) if vals else 0.0

    open_pnl    = safe_sum(active, "pnl_current")
    delta_total = safe_sum(active, "delta_current")
    max_loss_exp = safe_sum(active, "max_loss")
    n_active    = len(active)

    # RLZD por env
    rlzd_by_env = {}
    if not monthly.empty:
        for env, grp in monthly.groupby("environment"):
            rlzd_by_env[env] = round(grp["pnl_last"].sum(), 2)

    # Win rate (todos os trades com pnl_last)
    win_rate = None
    if not monthly.empty:
        total = len(monthly)
        wins  = monthly["won"].sum()
        win_rate = round(wins / total * 100, 1) if total > 0 else None

    # Alertas
    alerts = []
    for t in active:
        if t.get("alert_dte"):
            alerts.append(f"DTE CRÍTICO: {t['name']} com {t['dte_remaining']} DTE")
        if t.get("alert_stop"):
            alerts.append(f"STOP ATINGIDO: {t['name']} @ {t['pnl_pct_max']:.0f}% do Max Profit")
        if t.get("alert_profit_50"):
            alerts.append(f"50% TARGET: {t['name']} @ {t['pnl_pct_max']:.0f}% do Max Profit")

    return {
        "open_pnl_total":    open_pnl,
        "delta_total":       delta_total,
        "max_loss_exposed":  max_loss_exp,
        "n_active_trades":   n_active,
        "n_closed_trades":   len(inactive),
        "rlzd_by_env":       rlzd_by_env,
        "win_rate":          win_rate,
        "alerts":            alerts,
    }


# ─── Main export ─────────────────────────────────────────────────────────────

def run_export(xlsx_path: Path, gdrive_file_id: str | None = None, snapshot_only: bool = False) -> dict:
    today_str = date.today().strftime("%Y-%m-%d")

    # ── Download do Google Drive (se configurado) ──
    if gdrive_file_id:
        print(f"[export] Baixando do Google Drive (ID: {gdrive_file_id[:12]}...) ...")
        ok = download_from_gdrive(gdrive_file_id, xlsx_path)
        if ok:
            print(f"[export] Download OK -> {xlsx_path.name}")
        else:
            print(f"[export] Fallback: usando arquivo local {xlsx_path.name}")

    print(f"[export] Lendo {xlsx_path.name} ...")
    db_robots = load_db_robots(xlsx_path)
    db_cria   = load_db_cria(xlsx_path)

    print(f"[export] db_robots: {len(db_robots)} registros | db_cria: {len(db_cria)} trades")

    print("[export] Lendo TOTAL rows das abas visuais ...")
    sheet_summaries = read_sheet_summaries(xlsx_path)
    for env, months in sheet_summaries.items():
        for m in months:
            print(f"  {m['month']:15s} [{env:12s}]  OpenPnL={m['open_pnl']:+,.0f}  RLZD={m['rlzd']:+,.0f}  Delta={m['delta']:+.0f}")

    print("[export] Lendo PnLs individuais das abas visuais ...")
    individual_trade_pnls, sheet_to_trades, trade_visual_details, sheet_daily_pnls = read_individual_trade_pnls(xlsx_path)
    # Reverse map: trade_name -> sheet (first sheet wins if duplicated)
    trade_to_sheet: dict[str, str] = {}
    for sheet_name, names in sheet_to_trades.items():
        for n in names:
            trade_to_sheet.setdefault(n, sheet_name)

    print("[export] Lendo ForwardTest registry (opcional) ...")
    ft_strategies = load_ft_strategies(xlsx_path)
    print(f"[export] FT strategies (metadata): {len(ft_strategies)}")

    print("[export] Lendo status Closed das abas visuais ...")
    closed_from_sheets = {
        name
        for name, details in trade_visual_details.items()
        if "closed" in str(details.get("status") or "").lower()
    }
    print(f"[export] Trades fechados: {sorted(closed_from_sheets)}")

    trades   = build_trade_snapshot(db_robots, db_cria, closed_from_sheets)
    monthly  = build_monthly_summary(db_robots, db_cria)
    history  = pd.DataFrame() if snapshot_only else build_trade_history(db_robots, db_cria)

    # Attach `sheet` (visual month sheet, e.g. APR26) to each trade for filtering
    for t in trades:
        t["sheet"] = trade_to_sheet.get(t.get("name"))
        details = trade_visual_details.get(t.get("name"))
        if details:
            t["visual_pnl"] = details.get("pnl")
            t["dte_bucket"] = details.get("dte_bucket")
            t["visual_status"] = details.get("status")
            t["visual_open_date"] = details.get("open_date")
            t["visual_exp_date"] = details.get("exp_date")
            t["visual_dte_open"] = details.get("dte_open")
            t["visual_open_price"] = details.get("open_price")
            t["visual_last_pnl"] = details.get("last_pnl")
            t["visual_last_pnl_date"] = details.get("last_pnl_date")
            t["visual_last_pnl_weekday"] = details.get("last_pnl_weekday")
            t["visual_last_pnl_row"] = details.get("last_pnl_row")
            t["inferred_close_date"] = details.get("inferred_close_date")
            t["inferred_close_weekday"] = details.get("inferred_close_weekday")
            t["days_held"] = details.get("days_held")
            t["days_open_as_of_last_pnl"] = details.get("days_open_as_of_last_pnl")
            if not t.get("strikes"):
                t["strikes"] = details.get("strikes")
            if t.get("underlying") == "?":
                t["underlying"] = _infer_underlying_from_visual(t["name"], details.get("open_price"))

    # Some closed visual trades no longer have a db_robots current-state row.
    # Keep them in the snapshot so month filters, AI context, and future charts
    # reconcile to the visual sheet composition.
    existing_names = {t.get("name") for t in trades}
    for sheet_name, names in sheet_to_trades.items():
        env_norm = infer_visual_sheet_env(sheet_name)
        for name in names:
            if name in existing_names:
                continue
            pnl = individual_trade_pnls.get(name)
            details = trade_visual_details.get(name, {})
            is_active = name not in closed_from_sheets
            trades.append({
                "name": name,
                "environment_raw": sheet_name,
                "environment": env_norm,
                "underlying": _infer_underlying_from_visual(name, details.get("open_price")),
                "is_active": is_active,
                "last_update": None,
                "open_date": details.get("open_date"),
                "exp_date": details.get("exp_date"),
                "dte_open": details.get("dte_open"),
                "dte_remaining": None,
                "underlying_price_at_open": details.get("open_price"),
                "strikes": details.get("strikes"),
                "net_credit": None,
                "max_loss": None,
                "sd": None,
                "lw_be": None,
                "up_be": None,
                "pct_to_lw_be": None,
                "pct_to_up_be": None,
                "pnl_current": pnl,
                "pnl_pct_max": None,
                "delta_current": 0.0,
                "alert_dte": False,
                "alert_profit_50": False,
                "alert_stop": False,
                "tent_status": "unknown",
                "sheet": sheet_name,
                "source": "visual_sheet",
                "visual_pnl": pnl,
                "dte_bucket": details.get("dte_bucket"),
                "visual_status": details.get("status"),
                "visual_open_date": details.get("open_date"),
                "visual_exp_date": details.get("exp_date"),
                "visual_dte_open": details.get("dte_open"),
                "visual_open_price": details.get("open_price"),
                "visual_last_pnl": details.get("last_pnl"),
                "visual_last_pnl_date": details.get("last_pnl_date"),
                "visual_last_pnl_weekday": details.get("last_pnl_weekday"),
                "visual_last_pnl_row": details.get("last_pnl_row"),
                "inferred_close_date": details.get("inferred_close_date"),
                "inferred_close_weekday": details.get("inferred_close_weekday"),
                "days_held": details.get("days_held"),
                "days_open_as_of_last_pnl": details.get("days_open_as_of_last_pnl"),
            })
            existing_names.add(name)

    trades = sorted(trades, key=lambda t: (not t["is_active"], t["name"]))
    _attach_forward_daily_history(trades, db_robots)
    kpis = compute_portfolio_kpis(trades, monthly)

    # ── Snapshot JSON ──
    snapshot = {
        "generated_at":        datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source_file":         xlsx_path.name,
        "portfolio":           kpis,
        "sheet_summaries":     sheet_summaries,
        "individual_trade_pnls": individual_trade_pnls,
        "sheet_daily_pnls":    sheet_daily_pnls,
        "forwardtest_strategies": ft_strategies,
        "trades":              trades,
    }
    snap_path = REPORTS_DIR / f"trades_snapshot_{today_str}.json"
    with open(snap_path, "w", encoding="utf-8") as f:
        json.dump(snapshot, f, ensure_ascii=False, indent=2, default=str)
    print(f"[export] Snapshot -> {snap_path.name}")

    # -- Latest snapshot alias --
    latest_path = REPORTS_DIR / "trades_snapshot_latest.json"
    with open(latest_path, "w", encoding="utf-8") as f:
        json.dump(snapshot, f, ensure_ascii=False, indent=2, default=str)

    # -- Trade history parquet --
    if not snapshot_only and not history.empty:
        hist_path = REPORTS_DIR / "trade_history.parquet"
        history.to_parquet(hist_path, index=False)
        print(f"[export] History  -> {hist_path.name} ({len(history)} linhas)")

    # -- Monthly summary CSV --
    if not snapshot_only and not monthly.empty:
        csv_path = REPORTS_DIR / "monthly_summary.csv"
        try:
            monthly.to_csv(csv_path, index=False)
            print(f"[export] Monthly  -> {csv_path.name} ({len(monthly)} trades)")
        except PermissionError:
            print(f"[!] Monthly CSV bloqueado; snapshot ja foi atualizado. Feche o arquivo e rode novamente: {csv_path}")
    elif snapshot_only:
        print("[export] Snapshot-only: pulando parquet/CSV auxiliares")

    # -- Alertas --
    if kpis["alerts"]:
        print("\n[!] ALERTAS:")
        for a in kpis["alerts"]:
            print(f"   {a}")

    return snapshot


def get_latest_snapshot() -> dict | None:
    """Retorna o snapshot mais recente (usado por outros módulos)."""
    path = REPORTS_DIR / "trades_snapshot_latest.json"
    if not path.exists():
        # Fallback: procurar pelo mais recente datado
        candidates = sorted(REPORTS_DIR.glob("trades_snapshot_2*.json"), reverse=True)
        if not candidates:
            return None
        path = candidates[0]
    with open(path, encoding="utf-8") as f:
        return json.load(f)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Exporta OP Control Panel para JSON/parquet/CSV")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Caminho para o .xlsx")
    parser.add_argument(
        "--gdrive-id",
        default=os.environ.get("GDRIVE_FILE_ID", ""),
        help="ID do Google Sheet para download automatico (ou variavel GDRIVE_FILE_ID)",
    )
    parser.add_argument(
        "--snapshot-only",
        action="store_true",
        help="Gera apenas trades_snapshot_*.json e pula parquet/CSV auxiliares.",
    )
    args = parser.parse_args()

    xlsx = Path(args.xlsx)
    gdrive_id = args.gdrive_id.strip() or None

    if not gdrive_id and not xlsx.exists():
        print(f"[ERROR] Arquivo nao encontrado: {xlsx}")
        sys.exit(1)

    run_export(xlsx, gdrive_file_id=gdrive_id, snapshot_only=args.snapshot_only)
