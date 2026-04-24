#!/usr/bin/env python3
"""
Add an executive dashboard and AI export layer on top of the current OP Control Panel workbook.

The goal is to preserve the user's existing workbook structure and automations while adding:
- a shareable dashboard for Cristiano,
- a live helper sheet that normalizes active/closed trades from the current tabs,
- an AI export sheet with prompt-ready summaries,
- an AI workflow sheet with report templates.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent.parent
SOURCE = ROOT / "data" / "OP Control Panel.xlsx"
OUTPUT = ROOT / "data" / "OP Control Panel - Dashboard Pack.xlsx"

ADDON_SHEETS = ["Dashboard", "AI Export", "AI Workflow", "db_dashboard_helper"]

COLORS = {
    "navy": "0E2238",
    "navy_2": "17314F",
    "sky": "DCE9F6",
    "paper": "F7FAFC",
    "white": "FFFFFF",
    "text": "223344",
    "muted": "6C7B8A",
    "grid": "D7E2EB",
    "green": "178C5B",
    "green_soft": "DDF4E8",
    "red": "C4494F",
    "red_soft": "F8E0E3",
    "amber": "D2911F",
    "amber_soft": "FBECD0",
    "blue_soft": "DFECF9",
}

THIN = Side(style="thin", color=COLORS["grid"])
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def style_range(ws, rng: str, *, fill_color: str | None = None, border: Border | None = None,
                font: Font | None = None, alignment: Alignment | None = None) -> None:
    for row in ws[rng]:
        for cell in row:
            if fill_color:
                cell.fill = fill(fill_color)
            if border:
                cell.border = border
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment


def title_block(ws, title: str, subtitle: str, width_end: str = "N") -> None:
    ws.merge_cells(f"A1:{width_end}2")
    ws["A1"] = title
    ws["A1"].font = Font(name="Aptos Display", size=20, bold=True, color=COLORS["white"])
    ws["A1"].fill = fill(COLORS["navy"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"A3:{width_end}3")
    ws["A3"] = subtitle
    ws["A3"].font = Font(name="Aptos", size=10, color=COLORS["white"])
    ws["A3"].fill = fill(COLORS["navy_2"])
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")


def section_label(ws, row: int, label: str, end_col: str = "N") -> None:
    ws.merge_cells(f"A{row}:{end_col}{row}")
    ws[f"A{row}"] = label
    ws[f"A{row}"].font = Font(name="Aptos", size=11, bold=True, color=COLORS["navy"])
    ws[f"A{row}"].fill = fill(COLORS["sky"])
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")


def kpi_card(ws, label_range: str, value_range: str, label: str, formula: str,
             tone: str, number_format: str = "General") -> None:
    ws.merge_cells(label_range)
    tl = label_range.split(":")[0]
    ws[tl] = label
    ws[tl].font = Font(name="Aptos", size=10, bold=True, color=COLORS["muted"])
    ws[tl].fill = fill(COLORS[tone])
    ws[tl].alignment = Alignment(horizontal="left", vertical="center")
    style_range(ws, label_range, border=BOX)

    ws.merge_cells(value_range)
    vt = value_range.split(":")[0]
    ws[vt] = formula
    ws[vt].font = Font(name="Aptos Display", size=18, bold=True, color=COLORS["text"])
    ws[vt].fill = fill(COLORS[tone])
    ws[vt].alignment = Alignment(horizontal="left", vertical="center")
    ws[vt].number_format = number_format
    style_range(ws, value_range, border=BOX)


def header_row(ws, row: int, headers: list[str]) -> None:
    for idx, value in enumerate(headers, start=1):
        cell = ws.cell(row, idx, value)
        cell.font = Font(name="Aptos", size=10, bold=True, color=COLORS["white"])
        cell.fill = fill(COLORS["navy"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX


def data_style(ws, start_row: int, end_row: int, end_col: int) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=end_col):
        for cell in row:
            cell.border = BOX
            cell.font = Font(name="Aptos", size=10, color=COLORS["text"])
            cell.alignment = Alignment(vertical="center")


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def remove_existing_addon_sheets(wb) -> None:
    for name in ADDON_SHEETS:
        if name in wb.sheetnames:
            del wb[name]


def build_helper(ws) -> None:
    title_block(ws, "db_dashboard_helper", "Normalized live helper table built from the current workbook.", "Q")
    headers = [
        "source_sheet", "status", "trade_name", "optionstrat_url", "created_dt", "open_underlying",
        "dte_open", "expiry_dt", "days_to_expiry", "is_active", "current_pnl", "current_delta",
        "last_snapshot_ts", "net_credit", "max_loss", "contracts", "trade_name_norm",
    ]
    header_row(ws, 5, headers)

    formula = (
        '=LET('
        'mar,FILTER(HSTACK(ARRAYFORMULA(IF(TRANSPOSE(MAR26!$V$4:$ABV$4)<>"","MAR26","")),'
        'TRANSPOSE(MAR26!$V$2:$ABV$2),TRANSPOSE(MAR26!$V$3:$ABV$3),TRANSPOSE(MAR26!$V$4:$ABV$4)),'
        'TRANSPOSE(MAR26!$V$4:$ABV$4)<>""),'
        'apr,FILTER(HSTACK(ARRAYFORMULA(IF(TRANSPOSE(APR26!$V$5:$ADN$5)<>"","APR26","")),'
        'TRANSPOSE(APR26!$V$2:$ADN$2),TRANSPOSE(APR26!$V$4:$ADN$4),TRANSPOSE(APR26!$V$5:$ADN$5)),'
        'TRANSPOSE(APR26!$V$5:$ADN$5)<>""),'
        'jsmar,FILTER(HSTACK(ARRAYFORMULA(IF(TRANSPOSE(\'JS-FOR MAR26\'!$V$4:$AAQ$4)<>"","JS-FOR MAR26","")),'
        'TRANSPOSE(\'JS-FOR MAR26\'!$V$2:$AAQ$2),TRANSPOSE(\'JS-FOR MAR26\'!$V$3:$AAQ$3),TRANSPOSE(\'JS-FOR MAR26\'!$V$4:$AAQ$4)),'
        'TRANSPOSE(\'JS-FOR MAR26\'!$V$4:$AAQ$4)<>""),'
        'jsapr,FILTER(HSTACK(ARRAYFORMULA(IF(TRANSPOSE(\'JS APR26\'!$Z$4:$AAU$4)<>"","JS APR26","")),'
        'TRANSPOSE(\'JS APR26\'!$Z$2:$AAU$2),TRANSPOSE(\'JS APR26\'!$Z$3:$AAU$3),TRANSPOSE(\'JS APR26\'!$Z$4:$AAU$4)),'
        'TRANSPOSE(\'JS APR26\'!$Z$4:$AAU$4)<>""),'
        'fortr,FILTER(HSTACK(ARRAYFORMULA(IF(TRANSPOSE(\'FOR Trades\'!$V$4:$ZZ$4)<>"","FOR Trades","")),'
        'TRANSPOSE(\'FOR Trades\'!$V$2:$ZZ$2),TRANSPOSE(\'FOR Trades\'!$V$3:$ZZ$3),TRANSPOSE(\'FOR Trades\'!$V$4:$ZZ$4)),'
        'TRANSPOSE(\'FOR Trades\'!$V$4:$ZZ$4)<>""),'
        'all_raw,VSTACK(mar,apr,jsmar,jsapr,fortr),'
        'SORT(FILTER(all_raw,INDEX(all_raw,,4)<>""),3,TRUE)'
        ')'
    )
    ws["A6"] = formula

    ws["E6"] = '=ARRAYFORMULA(IF(D6:D="","",IFNA(VLOOKUP(D6:D,{db_cria!B:B,db_cria!A:A},2,FALSE),"")))'
    ws["F6"] = '=ARRAYFORMULA(IF(D6:D="","",IFNA(VLOOKUP(D6:D,{db_cria!B:B,db_cria!F:F},2,FALSE),"")))'
    ws["G6"] = '=ARRAYFORMULA(IF(D6:D="","",IFNA(VLOOKUP(D6:D,{db_cria!B:B,db_cria!C:C},2,FALSE),"")))'
    ws["H6"] = '=ARRAYFORMULA(IF(E6:E="","",E6:E+G6:G))'
    ws["I6"] = '=ARRAYFORMULA(IF(H6:H="","",H6:H-TODAY()))'
    ws["J6"] = '=ARRAYFORMULA(IF(B6:B="","",NOT(REGEXMATCH(UPPER(B6:B),"CLOSED"))))'
    ws["K6"] = '=ARRAYFORMULA(IF(D6:D="","",IFERROR(LOOKUP(ROW(D6:D),ROW(D6:D)/(D6:D<>""),LOOKUP(2,1/(db_robots!$F:$F=D6:D),db_robots!$D:$D)),"")))'
    ws["L6"] = '=ARRAYFORMULA(IF(D6:D="","",IFERROR(LOOKUP(ROW(D6:D),ROW(D6:D)/(D6:D<>""),LOOKUP(2,1/(db_robots!$F:$F=D6:D),db_robots!$E:$E)),"")))'
    ws["M6"] = '=ARRAYFORMULA(IF(D6:D="","",IFERROR(LOOKUP(ROW(D6:D),ROW(D6:D)/(D6:D<>""),LOOKUP(2,1/(db_robots!$F:$F=D6:D),db_robots!$A:$A)),"")))'
    ws["N6"] = '=ARRAYFORMULA(IF(D6:D="","",IFNA(VLOOKUP(D6:D,{db_cria!B:B,db_cria!I:I},2,FALSE),"")))'
    ws["O6"] = '=ARRAYFORMULA(IF(D6:D="","",IFNA(VLOOKUP(D6:D,{db_cria!B:B,db_cria!J:J},2,FALSE),"")))'
    ws["P6"] = '=ARRAYFORMULA(IF(D6:D="","",IFNA(VLOOKUP(D6:D,{db_cria!B:B,db_cria!K:K},2,FALSE),"")))'
    ws["Q6"] = '=ARRAYFORMULA(IF(C6:C="","",UPPER(TRIM(REGEXREPLACE(C6:C," - CLOSED","")))))'

    set_widths(ws, {
        "A": 16, "B": 12, "C": 34, "D": 34, "E": 16, "F": 14, "G": 10, "H": 14, "I": 12,
        "J": 10, "K": 12, "L": 12, "M": 18, "N": 12, "O": 12, "P": 28, "Q": 34,
    })
    ws.freeze_panes = "A6"


def build_dashboard(ws) -> None:
    title_block(ws, "Cristiano Dashboard", "Executive layer on top of the current control panel. Built to share.", "Y")
    ws.sheet_view.showGridLines = False
    set_widths(ws, {
        "A": 16, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16, "I": 16, "J": 16,
        "K": 16, "L": 16, "M": 16, "N": 16, "O": 16, "P": 16, "Q": 16, "R": 16, "S": 16, "T": 16,
        "U": 16, "V": 16, "W": 16, "X": 16, "Y": 16,
    })

    section_label(ws, 5, "Desk Pulse", "L")
    kpi_card(ws, "A6:C6", "A7:C8", "Active Trades", '=COUNTIF(db_dashboard_helper!J6:J,TRUE)', "blue_soft", "0")
    kpi_card(ws, "D6:F6", "D7:F8", "Open PnL", '=SUM(FILTER(db_dashboard_helper!K6:K,db_dashboard_helper!J6:J=TRUE))', "blue_soft", '"$"#,##0')
    kpi_card(ws, "G6:I6", "G7:I8", "Net Delta", '=SUM(FILTER(db_dashboard_helper!L6:L,db_dashboard_helper!J6:J=TRUE))', "amber_soft", "0")
    kpi_card(ws, "J6:L6", "J7:L8", "Next Expiry", '=MIN(FILTER(db_dashboard_helper!H6:H,db_dashboard_helper!J6:J=TRUE))', "amber_soft", "yyyy-mm-dd")

    kpi_card(ws, "A10:C10", "A11:C12", "Closed Trades", '=COUNTIF(db_dashboard_helper!J6:J,FALSE)', "green_soft", "0")
    kpi_card(ws, "D10:F10", "D11:F12", "Tracked URLs", '=COUNTA(FILTER(db_dashboard_helper!D6:D,db_dashboard_helper!D6:D<>""))', "blue_soft", "0")
    kpi_card(ws, "G10:I10", "G11:I12", "Snapshot Freshness", '=MAX(db_dashboard_helper!M6:M)', "blue_soft", "yyyy-mm-dd hh:mm")
    kpi_card(ws, "J10:L10", "J11:L12", "Avg Active DTE", '=AVERAGE(FILTER(db_dashboard_helper!I6:I,db_dashboard_helper!J6:J=TRUE))', "green_soft", "0.0")

    section_label(ws, 14, "Priority Feed", "H")
    header_row(ws, 15, ["Priority", "Trade", "Sheet", "Open PnL", "Delta", "DTE", "URL", "Action"])
    ws["A16"] = (
        '=SORT(FILTER({'
        'IF(ABS(db_dashboard_helper!L6:L)>=50,"HIGH",IF(ABS(db_dashboard_helper!L6:L)>=20,"MED","LOW")),'
        'db_dashboard_helper!C6:C,db_dashboard_helper!A6:A,db_dashboard_helper!K6:K,db_dashboard_helper!L6:L,'
        'db_dashboard_helper!I6:I,db_dashboard_helper!D6:D,'
        'IF(db_dashboard_helper!I6:I<=7,"Review expiry",IF(ABS(db_dashboard_helper!L6:L)>=50,"Review hedge","Monitor"))},'
        'db_dashboard_helper!J6:J=TRUE),1,TRUE,4,TRUE)'
    )
    data_style(ws, 16, 28, 8)

    section_label(ws, 30, "Active Trade Strip", "J")
    header_row(ws, 31, ["Trade", "Sheet", "Status", "Open Px", "Current PnL", "Delta", "Days To Exp", "Last Snapshot", "URL", "Contracts"])
    ws["A32"] = (
        '=SORT(FILTER({db_dashboard_helper!C6:C,db_dashboard_helper!A6:A,db_dashboard_helper!B6:B,db_dashboard_helper!F6:F,'
        'db_dashboard_helper!K6:K,db_dashboard_helper!L6:L,db_dashboard_helper!I6:I,db_dashboard_helper!M6:M,'
        'db_dashboard_helper!D6:D,db_dashboard_helper!P6:P},db_dashboard_helper!J6:J=TRUE),5,TRUE)'
    )
    data_style(ws, 32, 48, 10)

    # Helper metrics for charts
    ws["AA1"] = "metric"
    ws["AB1"] = "value"
    ws["AA2"] = "APR26"
    ws["AB2"] = '=COUNTIFS(db_dashboard_helper!A6:A,"APR26",db_dashboard_helper!J6:J,TRUE)'
    ws["AA3"] = "MAR26"
    ws["AB3"] = '=COUNTIFS(db_dashboard_helper!A6:A,"MAR26",db_dashboard_helper!J6:J,TRUE)'
    ws["AA4"] = "JS-FOR MAR26"
    ws["AB4"] = '=COUNTIFS(db_dashboard_helper!A6:A,"JS-FOR MAR26",db_dashboard_helper!J6:J,TRUE)'
    ws["AA5"] = "FOR Trades"
    ws["AB5"] = '=COUNTIFS(db_dashboard_helper!A6:A,"FOR Trades",db_dashboard_helper!J6:J,TRUE)'
    ws["AA6"] = "JS APR26"
    ws["AB6"] = '=COUNTIFS(db_dashboard_helper!A6:A,"JS APR26",db_dashboard_helper!J6:J,TRUE)'

    ws["AD1"] = "date"
    ws["AE1"] = "snapshot_rows"
    for r in range(2, 12):
        ws[f"AD{r}"] = f'=TODAY()-{12-r}'
        ws[f"AE{r}"] = f'=COUNTIF(db_robots!A:A,AD{r})'

    active_chart = BarChart()
    active_chart.title = "Active Trades by Sheet"
    active_chart.height = 6
    active_chart.width = 8
    active_chart.y_axis.title = "Trades"
    active_chart.x_axis.title = "Sheet"
    data = Reference(ws, min_col=28, min_row=1, max_row=6)
    cats = Reference(ws, min_col=27, min_row=2, max_row=6)
    active_chart.add_data(data, titles_from_data=True)
    active_chart.set_categories(cats)
    ws.add_chart(active_chart, "N6")

    sync_chart = LineChart()
    sync_chart.title = "Snapshot Volume (Last 10 Days)"
    sync_chart.height = 6
    sync_chart.width = 8
    sync_chart.y_axis.title = "Rows"
    sync_chart.x_axis.title = "Date"
    data = Reference(ws, min_col=31, min_row=1, max_row=11)
    cats = Reference(ws, min_col=30, min_row=2, max_row=11)
    sync_chart.add_data(data, titles_from_data=True)
    sync_chart.set_categories(cats)
    ws.add_chart(sync_chart, "N18")

    ws.conditional_formatting.add("A16:A28", FormulaRule(formula=['$A16="HIGH"'], fill=fill(COLORS["red_soft"])))
    ws.conditional_formatting.add("A16:A28", FormulaRule(formula=['$A16="MED"'], fill=fill(COLORS["amber_soft"])))
    ws.conditional_formatting.add("A16:A28", FormulaRule(formula=['$A16="LOW"'], fill=fill(COLORS["green_soft"])))


def build_ai_export(ws) -> None:
    title_block(ws, "AI Export", "Use this tab to send structured trade context to an AI and receive consistent reports.", "N")
    ws.sheet_view.showGridLines = False
    set_widths(ws, {c: 18 for c in "ABCDEFGHIJKLMN"})

    section_label(ws, 5, "Trade Selector", "H")
    ws["A6"] = "Selected URL"
    ws["B6"] = '=IFERROR(INDEX(FILTER(db_dashboard_helper!D6:D,db_dashboard_helper!J6:J=TRUE),1),"")'
    ws["A6"].font = Font(bold=True, color=COLORS["navy"])
    ws["B6"].fill = fill(COLORS["amber_soft"])
    ws["B6"].border = BOX

    detail_rows = [
        ("A8", "Trade Name", '=IFERROR(INDEX(FILTER(db_dashboard_helper!C6:C,db_dashboard_helper!D6:D=$B$6),1),"")'),
        ("A9", "Source Sheet", '=IFERROR(INDEX(FILTER(db_dashboard_helper!A6:A,db_dashboard_helper!D6:D=$B$6),1),"")'),
        ("A10", "Status", '=IFERROR(INDEX(FILTER(db_dashboard_helper!B6:B,db_dashboard_helper!D6:D=$B$6),1),"")'),
        ("A11", "Created Dt", '=IFERROR(INDEX(FILTER(db_dashboard_helper!E6:E,db_dashboard_helper!D6:D=$B$6),1),"")'),
        ("A12", "Open Underlying", '=IFERROR(INDEX(FILTER(db_dashboard_helper!F6:F,db_dashboard_helper!D6:D=$B$6),1),"")'),
        ("A13", "DTE Open", '=IFERROR(INDEX(FILTER(db_dashboard_helper!G6:G,db_dashboard_helper!D6:D=$B$6),1),"")'),
        ("A14", "Expiry Dt", '=IFERROR(INDEX(FILTER(db_dashboard_helper!H6:H,db_dashboard_helper!D6:D=$B$6),1),"")'),
        ("A15", "Current PnL", '=IFERROR(INDEX(FILTER(db_dashboard_helper!K6:K,db_dashboard_helper!D6:D=$B$6),1),"")'),
        ("A16", "Current Delta", '=IFERROR(INDEX(FILTER(db_dashboard_helper!L6:L,db_dashboard_helper!D6:D=$B$6),1),"")'),
        ("A17", "Net Credit", '=IFERROR(INDEX(FILTER(db_dashboard_helper!N6:N,db_dashboard_helper!D6:D=$B$6),1),"")'),
        ("A18", "Max Loss", '=IFERROR(INDEX(FILTER(db_dashboard_helper!O6:O,db_dashboard_helper!D6:D=$B$6),1),"")'),
        ("A19", "Contracts", '=IFERROR(INDEX(FILTER(db_dashboard_helper!P6:P,db_dashboard_helper!D6:D=$B$6),1),"")'),
    ]
    for coord, label, formula in detail_rows:
        ws[coord] = label
        ws[coord].font = Font(name="Aptos", size=10, bold=True, color=COLORS["navy"])
        value_coord = coord.replace("A", "B")
        ws[value_coord] = formula
        ws[value_coord].border = BOX

    section_label(ws, 21, "Prompt Builder", "N")
    ws.merge_cells("A22:N31")
    ws["A22"] = (
        '=TEXTJOIN(CHAR(10),TRUE,'
        '"You are reviewing one options trade from our desk.",'
        '"Trade name: "&B8,'
        '"Source sheet: "&B9,'
        '"Status: "&B10,'
        '"Created date: "&TEXT(B11,"yyyy-mm-dd"),'
        '"Open underlying: "&B12,'
        '"DTE at open: "&B13,'
        '"Expiry: "&TEXT(B14,"yyyy-mm-dd"),'
        '"Current PnL: "&B15,'
        '"Current Delta: "&B16,'
        '"Net credit: "&B17,'
        '"Max loss: "&B18,'
        '"Contracts: "&B19,'
        '"OptionStrat URL: "&$B$6,'
        '"Tasks: assess trade quality, structural asymmetry, risk placement, sizing quality, exit quality, and key leaks.",'
        '"Return: executive summary, quality score 0-10, strengths, leaks, specific corrections, and repeatability verdict."'
        ')'
    )
    ws["A22"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A22"].font = Font(name="Aptos", size=10, color=COLORS["text"])
    style_range(ws, "A22:N31", fill_color=COLORS["paper"], border=BOX)

    section_label(ws, 33, "Latest Snapshots", "N")
    header_row(ws, 34, ["Snapshot Dt", "Environment", "Strategy", "PnL", "Delta", "URL"])
    ws["A35"] = '=FILTER({db_robots!A:A,db_robots!B:B,db_robots!C:C,db_robots!D:D,db_robots!E:E,db_robots!F:F},db_robots!F:F=$B$6)'
    data_style(ws, 35, 60, 6)

    section_label(ws, 62, "AI Report Output", "N")
    ws.merge_cells("A63:N74")
    ws["A63"] = (
        "Paste the AI response here after reviewing the prompt above.\n\n"
        "Suggested report structure:\n"
        "1. Executive summary\n"
        "2. Quality score (0-10)\n"
        "3. Strengths\n"
        "4. Leaks / weaknesses\n"
        "5. Was the trade aligned with the setup intent?\n"
        "6. Was the sizing/risk acceptable?\n"
        "7. Actionable correction for next trade"
    )
    ws["A63"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A63"].font = Font(name="Aptos", size=10, color=COLORS["text"])
    style_range(ws, "A63:N74", fill_color=COLORS["paper"], border=BOX)


def build_ai_workflow(ws) -> None:
    title_block(ws, "AI Workflow", "Recommended AI reporting workflows for trade quality, leaks, and post-trade review.", "N")
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 4, "B": 22, "C": 88, "D": 20, "E": 18, "F": 18, "G": 18, "H": 18, "I": 18, "J": 18, "K": 18, "L": 18, "M": 18, "N": 18})

    sections = [
        ("B5", "Best Workflow"),
        ("C6", "1. Use `AI Export` to select one active or closed trade."),
        ("C7", "2. Copy the prompt builder block into your AI."),
        ("C8", "3. Paste the returned report into the `AI Report Output` area."),
        ("C9", "4. For weekly review, run the same process on the 3-5 most important trades."),
        ("B11", "Suggested Reports"),
        ("C12", "Daily open-trade report: which active trade has the weakest structure right now?"),
        ("C13", "Weekly leak report: what recurring mistakes appeared across the week's trades?"),
        ("C14", "Setup quality report: compare RJL vs Batman vs IC vs Call Bear + SP."),
        ("C15", "Exit quality report: did we close too early, too late, or correctly?"),
        ("C16", "Risk discipline report: was sizing appropriate relative to max loss and setup quality?"),
        ("B18", "What To Ask The AI"),
        ("C19", "Always ask for a quality score, repeatability verdict, leak list, and specific corrections."),
        ("C20", "Ask the AI to separate structural issues from execution issues."),
        ("C21", "Ask it to identify what made a good trade good, not only why bad trades failed."),
        ("B23", "Important Constraint"),
        ("C24", "For this desk, AI review should focus on macro index structures, 7DTE+ horizon, technical alignment, and risk asymmetry."),
    ]
    for coord, text in sections:
        ws[coord] = text
        if coord.startswith("B"):
            ws[coord].font = Font(name="Aptos", size=12, bold=True, color=COLORS["navy"])
        else:
            ws[coord].font = Font(name="Aptos", size=10, color=COLORS["text"])
            ws[coord].alignment = Alignment(wrap_text=True)


def add_validations(wb) -> None:
    ai = wb["AI Export"]
    dv = DataValidation(type="list", formula1="=FILTER(db_dashboard_helper!$D$6:$D,db_dashboard_helper!$D$6:$D<>'')", allow_blank=True)
    ai.add_data_validation(dv)
    dv.add("B6")


def add_file_note(ws) -> None:
    ws["AA3"] = "This add-on keeps the existing workbook structure."
    ws["AA3"].font = Font(name="Aptos", size=9, italic=True, color=COLORS["muted"])


def main() -> None:
    if not SOURCE.exists():
        raise SystemExit(f"Source workbook not found: {SOURCE}")

    wb = load_workbook(SOURCE)
    remove_existing_addon_sheets(wb)

    dashboard = wb.create_sheet("Dashboard", 0)
    ai_export = wb.create_sheet("AI Export", 1)
    ai_workflow = wb.create_sheet("AI Workflow", 2)
    helper = wb.create_sheet("db_dashboard_helper")

    build_dashboard(dashboard)
    build_ai_export(ai_export)
    build_ai_workflow(ai_workflow)
    build_helper(helper)
    add_validations(wb)
    add_file_note(dashboard)

    helper.sheet_state = "hidden"

    wb.save(OUTPUT)
    print(f"Dashboard pack generated: {OUTPUT}")


if __name__ == "__main__":
    main()
