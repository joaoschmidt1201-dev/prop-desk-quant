#!/usr/bin/env python3
"""
Builds a polished workbook template for the redesigned trading control panel.

Target:
- professional Google Sheets front-end
- normalized hidden data model
- executive dashboard for Cristiano
- ops-friendly structure for Joao and Make
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "data" / "control_panel" / "OP_Control_Panel_Ideal_Template.xlsx"


COLORS = {
    "navy": "10253F",
    "navy_2": "17314F",
    "slate": "60758C",
    "steel": "DDE6EF",
    "paper": "F6F8FB",
    "white": "FFFFFF",
    "green": "1F8A5B",
    "green_soft": "DDF4E8",
    "red": "C8474D",
    "red_soft": "F8E1E3",
    "amber": "D48C1F",
    "amber_soft": "FCEBCF",
    "blue_soft": "DCEBFA",
    "text": "203040",
    "muted": "6B7C93",
    "grid": "D9E2EC",
}


THIN = Side(style="thin", color=COLORS["grid"])
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def style_range(ws, cell_range: str, *, fill_color: str | None = None, font: Font | None = None,
                alignment: Alignment | None = None, border: Border | None = None) -> None:
    for row in ws[cell_range]:
        for cell in row:
            if fill_color:
                cell.fill = fill(fill_color)
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = border


def title_block(ws, title: str, subtitle: str) -> None:
    ws.merge_cells("A1:N2")
    ws["A1"] = title
    ws["A1"].font = Font(name="Aptos Display", size=20, bold=True, color=COLORS["white"])
    ws["A1"].fill = fill(COLORS["navy"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A3:N3")
    ws["A3"] = subtitle
    ws["A3"].font = Font(name="Aptos", size=10, color=COLORS["white"])
    ws["A3"].fill = fill(COLORS["navy_2"])
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")


def section_label(ws, row: int, label: str, span: str = "A:N") -> None:
    start, end = span.split(":")
    ws.merge_cells(f"{start}{row}:{end}{row}")
    cell = ws[f"{start}{row}"]
    cell.value = label
    cell.font = Font(name="Aptos", size=11, bold=True, color=COLORS["navy"])
    cell.fill = fill(COLORS["steel"])
    cell.alignment = Alignment(horizontal="left", vertical="center")


def kpi_card(ws, label_tl: str, label_br: str, value_tl: str, value_br: str,
             label: str, formula: str, tone: str = "blue_soft",
             number_format: str = "General") -> None:
    ws.merge_cells(f"{label_tl}:{label_br}")
    label_cell = ws[label_tl]
    label_cell.value = label
    label_cell.font = Font(name="Aptos", size=10, bold=True, color=COLORS["muted"])
    label_cell.fill = fill(COLORS[tone])
    label_cell.alignment = Alignment(horizontal="left", vertical="center")
    style_range(ws, f"{label_tl}:{label_br}", border=BOX)

    ws.merge_cells(f"{value_tl}:{value_br}")
    value_cell = ws[value_tl]
    value_cell.value = formula
    value_cell.font = Font(name="Aptos Display", size=18, bold=True, color=COLORS["text"])
    value_cell.fill = fill(COLORS[tone])
    value_cell.alignment = Alignment(horizontal="left", vertical="center")
    value_cell.number_format = number_format
    style_range(ws, f"{value_tl}:{value_br}", border=BOX)


def header_row(ws, row: int, headers: list[str]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row, col, header)
        cell.font = Font(name="Aptos", size=10, bold=True, color=COLORS["white"])
        cell.fill = fill(COLORS["navy"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX


def data_cell_style(ws, start_row: int, end_row: int, end_col: int) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=end_col):
        for cell in row:
            cell.border = BOX
            cell.font = Font(name="Aptos", size=10, color=COLORS["text"])
            cell.alignment = Alignment(vertical="center")


def configure_dimensions(ws, widths: dict[str, float], row_heights: dict[int, float] | None = None) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    if row_heights:
        for row, height in row_heights.items():
            ws.row_dimensions[row].height = height


def build_lookups(ws) -> None:
    title_block(ws, "Lookup Tables", "Hidden helper lists for validations and automation.")
    sections = {
        "A5": ("status", ["NEW", "PENDING", "ACTIVE", "CLOSED", "REVIEW", "ARCHIVED"]),
        "C5": ("book", ["CR_LIVE", "CR_FORWARD", "JS", "RESEARCH"]),
        "E5": ("owner", ["Cristiano", "Joao", "Shared"]),
        "G5": ("sleeve", ["Core", "Forward", "Tactical", "Research"]),
        "I5": ("underlying", ["SPX", "NDX", "RUT", "SPY", "QQQ", "GLD", "SLV", "BTC"]),
        "K5": ("setup_family", ["IC", "RJL", "Batman", "Broken Wing", "Call Bear + SP", "Put Half Bat"]),
        "M5": ("event_type", ["CREATE", "SNAPSHOT", "ADJUST", "CLOSE", "REVIEW", "ERROR"]),
        "O5": ("snapshot_slot", ["AM", "PM", "MANUAL"]),
        "Q5": ("ai_ready", ["YES", "NO"]),
    }
    for anchor, (label, values) in sections.items():
        ws[anchor] = label
        ws[anchor].font = Font(bold=True, color=COLORS["navy"])
        col = ws[anchor].column
        row = ws[anchor].row + 1
        for value in values:
            ws.cell(row, col, value)
            row += 1
    configure_dimensions(ws, {c: 16 for c in "ABCDEFGHIJKLMNOPQR"})


def build_registry(ws) -> None:
    title_block(ws, "db_trade_registry", "One row per trade. Master system of record.")
    headers = [
        "trade_id", "book", "owner", "sleeve", "status", "trade_name", "setup_family",
        "setup_variant", "underlying", "optionstrat_url", "open_dt", "expiry_dt",
        "dte_open", "dte_current", "contracts", "net_credit", "max_profit", "max_loss",
        "open_underlying", "soll_low_be", "soll_up_be", "ist_low_be", "ist_up_be",
        "current_pnl", "realized_pnl", "current_delta", "tags", "month_bucket", "notes",
        "ai_ready", "last_snapshot_ts",
    ]
    header_row(ws, 5, headers)

    sample_rows = [
        [
            "CR-20260413-RUT-RJL42", "CR_LIVE", "Cristiano", "Core", "ACTIVE", "T45 RUT RJL42",
            "RJL", "42D", "RUT", "https://optionstrat.com/7UbmrRBBdfwE", "2026-04-13", "2026-05-22",
            39, '=IF(L6="","",L6-TODAY())', "2400P | 2450P | 2750C", 11850, 11850, 3150, 2645,
            2410.5, 2789.5, 2410.5, 2789.5,
            '=IFERROR(LOOKUP(2,1/(db_snapshots!B:B=A6),db_snapshots!D:D),"")',
            '=IF(E6="CLOSED",X6,"")',
            '=IFERROR(LOOKUP(2,1/(db_snapshots!B:B=A6),db_snapshots!F:F),"")',
            "RUT, 42D, tactical", '=TEXT(K6,"yyyy-mm")', "Current Cristiano live trade", "YES",
            '=IFERROR(LOOKUP(2,1/(db_snapshots!B:B=A6),db_snapshots!A:A),"")',
        ],
        [
            "CR-20260406-RUT-CBSP28", "CR_LIVE", "Cristiano", "Tactical", "CLOSED", "T42 RUT CALL BEAR + SP 28",
            "Call Bear + SP", "28D", "RUT", "https://optionstrat.com/8X3B4avk9c7A", "2026-04-06", "2026-05-01",
            25, '=IF(L7="","",L7-TODAY())', "2400P | 2550C | 2700C", 14500, 14500, 225500, 2537,
            2255, 2622.5, 2255, 2622.5,
            '=IFERROR(LOOKUP(2,1/(db_snapshots!B:B=A7),db_snapshots!D:D),"")',
            3900,
            '=IFERROR(LOOKUP(2,1/(db_snapshots!B:B=A7),db_snapshots!F:F),"")',
            "RUT, 28D, closed", '=TEXT(K7,"yyyy-mm")', "Closed tactical trade", "YES",
            '=IFERROR(LOOKUP(2,1/(db_snapshots!B:B=A7),db_snapshots!A:A),"")',
        ],
        [
            "FOR-20260317-SPX-HYBBAT42", "CR_FORWARD", "Cristiano", "Forward", "ACTIVE", "FOR03 SPX HYB BAT42",
            "Batman", "42D", "SPX", "https://optionstrat.com/rYW6GRjv4Sa8", "2026-03-17", "2026-04-24",
            38, '=IF(L8="","",L8-TODAY())', "6340P | 6480P | 6540P", 12615, 12615, 12615, 6723,
            6462, 7002, 6340, 6990,
            '=IFERROR(LOOKUP(2,1/(db_snapshots!B:B=A8),db_snapshots!D:D),"")',
            '=IF(E8="CLOSED",X8,"")',
            '=IFERROR(LOOKUP(2,1/(db_snapshots!B:B=A8),db_snapshots!F:F),"")',
            "SPX, forward", '=TEXT(K8,"yyyy-mm")', "Forward sleeve active trade", "YES",
            '=IFERROR(LOOKUP(2,1/(db_snapshots!B:B=A8),db_snapshots!A:A),"")',
        ],
        [
            "JS-20260323-GLD-PHB14", "JS", "Joao", "Research", "REVIEW", "GLD PUT-HALF BAT14",
            "Put Half Bat", "14D", "GLD", "https://optionstrat.com/OMi7M8cs9qKA", "2026-03-23", "2026-04-17",
            30, '=IF(L9="","",L9-TODAY())', "355P | 380P | 385P", 2580, 2580, 13420, 413,
            371.78, "", 371.78, "",
            '=IFERROR(LOOKUP(2,1/(db_snapshots!B:B=A9),db_snapshots!D:D),"")',
            '=IF(E9="CLOSED",X9,"")',
            '=IFERROR(LOOKUP(2,1/(db_snapshots!B:B=A9),db_snapshots!F:F),"")',
            "GLD, review", '=TEXT(K9,"yyyy-mm")', "Research template row", "NO",
            '=IFERROR(LOOKUP(2,1/(db_snapshots!B:B=A9),db_snapshots!A:A),"")',
        ],
    ]

    for row_idx, values in enumerate(sample_rows, start=6):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)

    configure_dimensions(
        ws,
        {
            "A": 24, "B": 14, "C": 14, "D": 12, "E": 12, "F": 28, "G": 18, "H": 16,
            "I": 12, "J": 34, "K": 12, "L": 12, "M": 10, "N": 10, "O": 28, "P": 12,
            "Q": 12, "R": 12, "S": 14, "T": 12, "U": 12, "V": 12, "W": 12, "X": 12,
            "Y": 12, "Z": 12, "AA": 20, "AB": 12, "AC": 28, "AD": 10, "AE": 18,
        },
        {5: 28},
    )
    data_cell_style(ws, 6, 1200, len(headers))
    ws.freeze_panes = "A6"

    green_fill = fill(COLORS["green_soft"])
    red_fill = fill(COLORS["red_soft"])
    amber_fill = fill(COLORS["amber_soft"])
    ws.conditional_formatting.add("E6:E1200", FormulaRule(formula=['$E6="ACTIVE"'], fill=green_fill))
    ws.conditional_formatting.add("E6:E1200", FormulaRule(formula=['$E6="CLOSED"'], fill=red_fill))
    ws.conditional_formatting.add("E6:E1200", FormulaRule(formula=['$E6="REVIEW"'], fill=amber_fill))


def build_trade_legs(ws) -> None:
    title_block(ws, "db_trade_legs", "One row per leg.")
    headers = ["trade_id", "leg_no", "action", "qty", "option_type", "strike", "expiry_dt", "delta_open"]
    header_row(ws, 5, headers)
    rows = [
        ["CR-20260413-RUT-RJL42", 1, "SELL", 1, "PUT", 2450, "2026-05-22", -0.22],
        ["CR-20260413-RUT-RJL42", 2, "BUY", 1, "PUT", 2400, "2026-05-22", -0.12],
        ["CR-20260413-RUT-RJL42", 3, "SELL", 1, "CALL", 2750, "2026-05-22", 0.18],
        ["FOR-20260317-SPX-HYBBAT42", 1, "BUY", 1, "PUT", 6340, "2026-04-24", -0.08],
        ["FOR-20260317-SPX-HYBBAT42", 2, "SELL", 1, "PUT", 6480, "2026-04-24", -0.15],
        ["FOR-20260317-SPX-HYBBAT42", 3, "SELL", 1, "CALL", 6540, "2026-04-24", 0.14],
    ]
    for row_idx, values in enumerate(rows, start=6):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)
    configure_dimensions(ws, {c: 18 for c in "ABCDEFGH"})
    data_cell_style(ws, 6, 500, len(headers))
    ws.freeze_panes = "A6"


def build_snapshots(ws) -> None:
    title_block(ws, "db_snapshots", "One row per snapshot pull.")
    headers = [
        "snapshot_ts", "trade_id", "snapshot_slot", "pnl_open", "pnl_realized", "delta",
        "underlying_price", "vix", "regime_tag", "dist_soll_low_pct", "dist_soll_up_pct",
        "inside_tent_flag", "note",
    ]
    header_row(ws, 5, headers)
    rows = [
        ["2026-04-13 10:00", "CR-20260413-RUT-RJL42", "AM", -250, 0, -42, 2645, 24.1, "Balanced", 0.097, -0.037, "YES", ""],
        ["2026-04-13 16:00", "CR-20260413-RUT-RJL42", "PM", -180, 0, -39, 2652, 24.4, "Balanced", 0.100, -0.034, "YES", ""],
        ["2026-04-14 10:00", "CR-20260413-RUT-RJL42", "AM", -950, 0, -61, 2618, 25.8, "Stress", 0.086, -0.061, "YES", ""],
        ["2026-04-06 10:00", "CR-20260406-RUT-CBSP28", "AM", 1200, 0, -18, 2537, 25.1, "Trend Up", 0.111, 0.034, "YES", ""],
        ["2026-04-10 16:00", "CR-20260406-RUT-CBSP28", "PM", 3900, 3900, 0, 2571, 22.2, "Take Profit", 0.140, 0.052, "YES", "Closed"],
        ["2026-03-17 10:00", "FOR-20260317-SPX-HYBBAT42", "AM", 0, 0, -19, 6723, 19.9, "Open", 0.039, 0.041, "YES", ""],
        ["2026-03-18 16:00", "FOR-20260317-SPX-HYBBAT42", "PM", 1835, 0, -22, 6624.7, 20.8, "Pullback", 0.025, 0.055, "YES", ""],
        ["2026-03-23 10:00", "JS-20260323-GLD-PHB14", "AM", 2580, 0, -8, 413, 19.3, "Commodity", 0.020, 0.000, "YES", ""],
    ]
    for row_idx, values in enumerate(rows, start=6):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)
    configure_dimensions(ws, {"A": 18, "B": 24, "C": 12, "D": 12, "E": 12, "F": 10,
                              "G": 14, "H": 10, "I": 14, "J": 14, "K": 14, "L": 12, "M": 26})
    data_cell_style(ws, 6, 5000, len(headers))
    ws.freeze_panes = "A6"


def build_events(ws) -> None:
    title_block(ws, "db_events", "One row per state change or review event.")
    headers = ["event_id", "trade_id", "event_ts", "event_type", "actor", "field_name", "old_value", "new_value", "note"]
    header_row(ws, 5, headers)
    rows = [
        ["EVT-0001", "CR-20260413-RUT-RJL42", "2026-04-13 09:32", "CREATE", "Make", "status", "NEW", "ACTIVE", "Trade inserted from scraper"],
        ["EVT-0002", "CR-20260406-RUT-CBSP28", "2026-04-10 15:57", "CLOSE", "Joao", "status", "ACTIVE", "CLOSED", "Target reached"],
        ["EVT-0003", "JS-20260323-GLD-PHB14", "2026-03-30 11:00", "REVIEW", "Joao", "review_status", "", "REVIEW", "Flag for AI leak review"],
    ]
    for row_idx, values in enumerate(rows, start=6):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)
    configure_dimensions(ws, {"A": 12, "B": 24, "C": 18, "D": 14, "E": 12, "F": 16, "G": 14, "H": 14, "I": 30})
    data_cell_style(ws, 6, 2000, len(headers))
    ws.freeze_panes = "A6"


def build_import_queue(ws) -> None:
    title_block(ws, "db_import_queue", "Queue of newly detected trades before registry write.")
    headers = ["detected_ts", "source_book", "source_sheet", "source_column_group", "raw_status",
               "raw_trade_name", "optionstrat_url", "processed_flag", "processed_ts", "target_trade_id", "error_message"]
    header_row(ws, 5, headers)
    rows = [
        ["2026-04-21 10:01", "CR_LIVE", "APR26", "BQ", "NEW", "T46 RUT RJL7", "https://optionstrat.com/PzvlAjGeuUaS", "YES", "2026-04-21 10:02", "CR-20260421-RUT-RJL7", ""],
        ["2026-04-22 10:01", "CR_FORWARD", "FOR Trades", "AN", "NEW", "", "", "NO", "", "", "Missing URL"],
    ]
    for row_idx, values in enumerate(rows, start=6):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)
    configure_dimensions(ws, {"A": 18, "B": 14, "C": 16, "D": 16, "E": 12, "F": 28, "G": 34, "H": 12, "I": 18, "J": 24, "K": 24})
    data_cell_style(ws, 6, 2000, len(headers))
    ws.freeze_panes = "A6"


def build_make_control(ws) -> None:
    title_block(ws, "db_make_control", "Central control values for Make scenarios.")
    headers = ["control_key", "control_value", "last_update_ts", "notes"]
    header_row(ws, 5, headers)
    rows = [
        ["last_new_trade_scan_ts", "2026-04-23 09:30", "2026-04-23 09:30", "Latest scan successful"],
        ["last_snapshot_sync_ts", "2026-04-23 16:05", "2026-04-23 16:05", "PM sync finished"],
        ["active_trade_count", 4, "2026-04-23 16:05", "Derived count"],
        ["snapshot_failures_last_7d", 0, "2026-04-23 16:05", "Healthy"],
        ["last_error_message", "", "2026-04-23 16:05", ""],
    ]
    for row_idx, values in enumerate(rows, start=6):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)
    configure_dimensions(ws, {"A": 26, "B": 22, "C": 18, "D": 34})
    data_cell_style(ws, 6, 100, len(headers))


def build_dashboard(ws) -> None:
    title_block(ws, "Desk Dashboard", "Executive view for Cristiano. Operationally safe, visually clean, share-ready.")
    ws.sheet_view.showGridLines = False
    configure_dimensions(
        ws,
        {
            "A": 16, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16,
            "H": 16, "I": 16, "J": 16, "K": 16, "L": 16, "M": 16, "N": 16,
            "T": 16, "U": 16, "V": 16, "W": 16, "X": 16, "Y": 16,
        },
        {1: 24, 2: 24, 3: 20, 6: 36, 10: 36},
    )

    section_label(ws, 5, "Desk Pulse")
    kpi_card(ws, "A6", "C6", "A7", "C8", "Active Trades", "=U2", "blue_soft", "0")
    kpi_card(ws, "D6", "F6", "D7", "F8", "Open PnL", "=U3", "blue_soft", '"$"#,##0')
    kpi_card(ws, "G6", "I6", "G7", "I8", "Realized MTD", "=U4", "green_soft", '"$"#,##0')
    kpi_card(ws, "J6", "L6", "J7", "L8", "Win Rate 90D", "=U5", "green_soft", "0.0%")

    kpi_card(ws, "A10", "C10", "A11", "C12", "Net Delta", "=U6", "amber_soft", "0")
    kpi_card(ws, "D10", "F10", "D11", "F12", "Open Risk", "=U7", "red_soft", '"$"#,##0')
    kpi_card(ws, "G10", "I10", "G11", "I12", "Next Expiry", "=U8", "amber_soft", "yyyy-mm-dd")
    kpi_card(ws, "J10", "L10", "J11", "L12", "Data Freshness", "=U9", "blue_soft", "yyyy-mm-dd hh:mm")

    section_label(ws, 14, "Attention Feed")
    header_row(ws, 15, ["Priority", "Trade", "Book", "Pnl", "Delta", "Expiry", "Action", "Note"])
    attention_rows = [
        ["HIGH", "T45 RUT RJL42", "CR_LIVE", '=IFERROR(LOOKUP(2,1/(db_trade_registry!A:A="CR-20260413-RUT-RJL42"),db_trade_registry!X:X),"")', -61, "2026-05-22", "Watch delta drift", "RUT tactical live trade"],
        ["MED", "FOR03 SPX HYB BAT42", "CR_FORWARD", '=IFERROR(LOOKUP(2,1/(db_trade_registry!A:A="FOR-20260317-SPX-HYBBAT42"),db_trade_registry!X:X),"")', -22, "2026-04-24", "Check BE distance", "Forward sleeve active trade"],
    ]
    for r_idx, values in enumerate(attention_rows, start=16):
        for c_idx, value in enumerate(values, start=1):
            ws.cell(r_idx, c_idx, value)
    data_cell_style(ws, 16, 20, 8)

    section_label(ws, 22, "Charts")
    ws["A23"] = "PnL by month"
    ws["H23"] = "Active trades by book"
    ws["A23"].font = Font(bold=True, color=COLORS["navy"])
    ws["H23"].font = Font(bold=True, color=COLORS["navy"])

    # Hidden helper metrics on the dashboard itself for chart references.
    helper = {
        "T1": "metric", "U1": "value",
        "T2": "active_trades", "U2": '=COUNTIFS(db_trade_registry!E:E,"ACTIVE")',
        "T3": "open_pnl", "U3": '=SUMIFS(db_trade_registry!X:X,db_trade_registry!E:E,"ACTIVE")',
        "T4": "realized_mtd", "U4": '=SUMIFS(db_trade_registry!Y:Y,db_trade_registry!AB:AB,TEXT(TODAY(),"yyyy-mm"))',
        "T5": "win_rate_90d", "U5": '=IFERROR(COUNTIFS(db_trade_registry!E:E,"CLOSED",db_trade_registry!Y:Y,">0")/COUNTIFS(db_trade_registry!E:E,"CLOSED"),0)',
        "T6": "net_delta", "U6": '=SUMIFS(db_trade_registry!Z:Z,db_trade_registry!E:E,"ACTIVE")',
        "T7": "open_risk", "U7": '=SUMIFS(db_trade_registry!R:R,db_trade_registry!E:E,"ACTIVE")',
        "T8": "next_expiry", "U8": '=MINIFS(db_trade_registry!L:L,db_trade_registry!E:E,"ACTIVE")',
        "T9": "freshness", "U9": '=MAX(db_snapshots!A:A)',
        "W1": "month", "X1": "realized_pnl",
        "W2": "2026-01", "X2": '=SUMIFS(db_trade_registry!Y:Y,db_trade_registry!AB:AB,W2)',
        "W3": "2026-02", "X3": '=SUMIFS(db_trade_registry!Y:Y,db_trade_registry!AB:AB,W3)',
        "W4": "2026-03", "X4": '=SUMIFS(db_trade_registry!Y:Y,db_trade_registry!AB:AB,W4)',
        "W5": "2026-04", "X5": '=SUMIFS(db_trade_registry!Y:Y,db_trade_registry!AB:AB,W5)',
        "Z1": "book", "AA1": "active_count",
        "Z2": "CR_LIVE", "AA2": '=COUNTIFS(db_trade_registry!B:B,Z2,db_trade_registry!E:E,"ACTIVE")',
        "Z3": "CR_FORWARD", "AA3": '=COUNTIFS(db_trade_registry!B:B,Z3,db_trade_registry!E:E,"ACTIVE")',
        "Z4": "JS", "AA4": '=COUNTIFS(db_trade_registry!B:B,Z4,db_trade_registry!E:E,"ACTIVE")',
    }
    for coord, value in helper.items():
        ws[coord] = value

    pnl_chart = LineChart()
    pnl_chart.title = "Realized PnL by Month"
    pnl_chart.height = 6
    pnl_chart.width = 8
    pnl_chart.y_axis.title = "USD"
    pnl_chart.x_axis.title = "Month"
    data = Reference(ws, min_col=24, min_row=1, max_row=5)
    cats = Reference(ws, min_col=23, min_row=2, max_row=5)
    pnl_chart.add_data(data, titles_from_data=True)
    pnl_chart.set_categories(cats)
    ws.add_chart(pnl_chart, "A24")

    book_chart = BarChart()
    book_chart.title = "Active Trades by Book"
    book_chart.height = 6
    book_chart.width = 8
    book_chart.y_axis.title = "Trades"
    book_chart.x_axis.title = "Book"
    data = Reference(ws, min_col=27, min_row=1, max_row=4)
    cats = Reference(ws, min_col=26, min_row=2, max_row=4)
    book_chart.add_data(data, titles_from_data=True)
    book_chart.set_categories(cats)
    ws.add_chart(book_chart, "H24")


def build_monitor(ws) -> None:
    title_block(ws, "Trade Monitor", "Primary operating grid for all active trades.")
    headers = [
        "trade_id", "book", "status", "trade_name", "setup_family", "underlying", "open_dt",
        "expiry_dt", "dte_current", "net_credit", "max_profit", "max_loss", "current_pnl",
        "realized_pnl", "current_delta", "last_snapshot_ts", "optionstrat_url",
    ]
    header_row(ws, 5, headers)
    for idx, width in enumerate([22, 14, 12, 28, 16, 12, 12, 12, 10, 12, 12, 12, 12, 12, 12, 18, 34], start=1):
        ws.column_dimensions[chr(64 + idx)].width = width if idx <= 26 else 16
    query = (
        '=QUERY(db_trade_registry!A1:AE,'
        "\"select A,B,E,F,G,I,K,L,N,P,Q,R,X,Y,Z,AE,J where E = 'ACTIVE'\",1)"
    )
    ws["A6"] = query
    ws["A3"] = "Google Sheets note: this tab becomes a live filtered view after upload."
    ws["A3"].font = Font(italic=True, color=COLORS["muted"])
    ws.freeze_panes = "A6"


def build_active(ws) -> None:
    title_block(ws, "Active Trades", "Clean viewer tab for Cristiano.")
    headers = ["trade_id", "book", "trade_name", "underlying", "expiry_dt", "current_pnl", "current_delta", "optionstrat_url"]
    header_row(ws, 5, headers)
    query = (
        '=QUERY(db_trade_registry!A1:AE,'
        "\"select A,B,F,I,L,X,Z,J where E = 'ACTIVE'\",1)"
    )
    ws["A6"] = query
    configure_dimensions(ws, {"A": 24, "B": 14, "C": 30, "D": 12, "E": 12, "F": 12, "G": 12, "H": 36})
    ws.freeze_panes = "A6"


def build_closed(ws) -> None:
    title_block(ws, "Closed Trades", "Post-trade review surface with room for leak detection.")
    headers = [
        "trade_id", "book", "trade_name", "underlying", "open_dt", "expiry_dt",
        "realized_pnl", "close_reason", "leak_tag", "strength_tag", "review_status",
    ]
    header_row(ws, 5, headers)
    query = (
        '=QUERY(db_trade_registry!A1:AE,'
        "\"select A,B,F,I,K,L,Y,AC,AA,AA,E where E = 'CLOSED'\",1)"
    )
    ws["A6"] = query
    configure_dimensions(ws, {"A": 24, "B": 14, "C": 30, "D": 12, "E": 12, "F": 12, "G": 12, "H": 18, "I": 18, "J": 18, "K": 14})
    ws.freeze_panes = "A6"


def build_performance(ws) -> None:
    title_block(ws, "Performance", "Historical analytics and setup quality.")
    ws.sheet_view.showGridLines = False
    configure_dimensions(ws, {c: 16 for c in "ABCDEFGHIJKLMNQRSTUVWX"})

    section_label(ws, 5, "Monthly Performance")
    header_row(ws, 6, ["Month", "Realized PnL"])
    months = ["2026-01", "2026-02", "2026-03", "2026-04"]
    for idx, month in enumerate(months, start=7):
        ws.cell(idx, 1, month)
        ws.cell(idx, 2, f'=SUMIFS(db_trade_registry!Y:Y,db_trade_registry!AB:AB,A{idx})')
    data_cell_style(ws, 7, 10, 2)

    section_label(ws, 12, "Setup Performance")
    header_row(ws, 13, ["Setup Family", "Closed Trades", "Total Realized", "Avg Realized"])
    setups = ["RJL", "Batman", "Call Bear + SP", "Put Half Bat"]
    for idx, setup in enumerate(setups, start=14):
        ws.cell(idx, 1, setup)
        ws.cell(idx, 2, f'=COUNTIFS(db_trade_registry!G:G,A{idx},db_trade_registry!E:E,"CLOSED")')
        ws.cell(idx, 3, f'=SUMIFS(db_trade_registry!Y:Y,db_trade_registry!G:G,A{idx},db_trade_registry!E:E,"CLOSED")')
        ws.cell(idx, 4, f'=IFERROR(C{idx}/B{idx},0)')
    data_cell_style(ws, 14, 17, 4)

    pnl_chart = LineChart()
    pnl_chart.title = "Monthly Realized PnL"
    pnl_chart.height = 6
    pnl_chart.width = 8
    data = Reference(ws, min_col=2, min_row=6, max_row=10)
    cats = Reference(ws, min_col=1, min_row=7, max_row=10)
    pnl_chart.add_data(data, titles_from_data=True)
    pnl_chart.set_categories(cats)
    ws.add_chart(pnl_chart, "F6")

    setup_chart = BarChart()
    setup_chart.title = "Total Realized by Setup"
    setup_chart.height = 6
    setup_chart.width = 8
    data = Reference(ws, min_col=3, min_row=13, max_row=17)
    cats = Reference(ws, min_col=1, min_row=14, max_row=17)
    setup_chart.add_data(data, titles_from_data=True)
    setup_chart.set_categories(cats)
    ws.add_chart(setup_chart, "F14")


def build_ai_export(ws) -> None:
    title_block(ws, "AI Export", "Controlled export surface for review prompts, leak analysis, and structured datasets.")
    configure_dimensions(ws, {"A": 18, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18, "I": 18, "J": 18})
    section_label(ws, 5, "Selected Trade")
    ws["A6"] = "Select trade_id"
    ws["B6"] = "CR-20260413-RUT-RJL42"
    ws["A6"].font = Font(bold=True, color=COLORS["navy"])
    ws["B6"].fill = fill(COLORS["amber_soft"])

    labels = [
        ("A8", "Trade Name", '=IFERROR(INDEX(db_trade_registry!F:F,MATCH($B$6,db_trade_registry!A:A,0)),"")'),
        ("A9", "Underlying", '=IFERROR(INDEX(db_trade_registry!I:I,MATCH($B$6,db_trade_registry!A:A,0)),"")'),
        ("A10", "Setup", '=IFERROR(INDEX(db_trade_registry!G:G,MATCH($B$6,db_trade_registry!A:A,0)),"")'),
        ("A11", "Status", '=IFERROR(INDEX(db_trade_registry!E:E,MATCH($B$6,db_trade_registry!A:A,0)),"")'),
        ("A12", "Open Dt", '=IFERROR(INDEX(db_trade_registry!K:K,MATCH($B$6,db_trade_registry!A:A,0)),"")'),
        ("A13", "Expiry", '=IFERROR(INDEX(db_trade_registry!L:L,MATCH($B$6,db_trade_registry!A:A,0)),"")'),
        ("A14", "Current PnL", '=IFERROR(INDEX(db_trade_registry!X:X,MATCH($B$6,db_trade_registry!A:A,0)),"")'),
        ("A15", "Current Delta", '=IFERROR(INDEX(db_trade_registry!Z:Z,MATCH($B$6,db_trade_registry!A:A,0)),"")'),
        ("A16", "URL", '=IFERROR(INDEX(db_trade_registry!J:J,MATCH($B$6,db_trade_registry!A:A,0)),"")'),
    ]
    for anchor, label, formula in labels:
        ws[anchor] = label
        ws[anchor].font = Font(bold=True, color=COLORS["navy"])
        ws[anchor.replace("A", "B")] = formula

    section_label(ws, 18, "Prompt Builder")
    ws.merge_cells("A19:J26")
    ws["A19"] = (
        '=TEXTJOIN(CHAR(10),TRUE,'
        '"Review this options trade for structural quality and leaks.",'
        '"trade_id: "&$B$6,'
        '"trade_name: "&B8,'
        '"underlying: "&B9,'
        '"setup: "&B10,'
        '"status: "&B11,'
        '"open_dt: "&TEXT(B12,"yyyy-mm-dd"),'
        '"expiry_dt: "&TEXT(B13,"yyyy-mm-dd"),'
        '"current_pnl: "&B14,'
        '"current_delta: "&B15,'
        '"optionstrat_url: "&B16,'
        '"Tasks: identify strengths, leaks, exit quality, sizing quality, and repeatability.")'
    )
    ws["A19"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A19"].font = Font(name="Aptos", size=10)
    style_range(ws, "A19:J26", fill_color=COLORS["paper"], border=BOX)


def build_playbook(ws) -> None:
    title_block(ws, "Desk Playbook", "Operating doctrine for the macro options desk.")
    configure_dimensions(ws, {"A": 4, "B": 24, "C": 80})
    rows = [
        ("B5", "Desk Rules"),
        ("C6", "1. Trade macro index products, sector ETFs, and macro commodities. Avoid single-name idiosyncratic risk."),
        ("C7", "2. Minimum operating horizon: 7 DTE. The core book remains 28-42 DTE."),
        ("C8", "3. Every premium sale must agree with chart structure and technical regime."),
        ("C9", "4. GEX and technical structure must align before risk is added."),
        ("B11", "Daily Rhythm"),
        ("C12", "AM: refresh active trade snapshots, review delta drift, confirm nearest expiries."),
        ("C13", "PM: refresh snapshots again, update action feed, flag trades for review."),
        ("B15", "Review Discipline"),
        ("C16", "Every closed trade should receive a close reason, leak tag, and strength tag."),
        ("C17", "AI review should be used after the structured review fields are filled."),
    ]
    for coord, value in rows:
        ws[coord] = value
        if coord.startswith("B"):
            ws[coord].font = Font(name="Aptos", size=12, bold=True, color=COLORS["navy"])
        else:
            ws[coord].font = Font(name="Aptos", size=11, color=COLORS["text"])
            ws[coord].alignment = Alignment(wrap_text=True)


def build_ops_control(ws) -> None:
    title_block(ws, "Ops Control", "Joao-facing operating console for data quality and migration.")
    configure_dimensions(ws, {"A": 24, "B": 44, "C": 18, "D": 18})
    header_row(ws, 5, ["Module", "Description", "Owner", "Status"])
    rows = [
        ["Registry", "Master trade registry. One row per trade.", "Joao", "READY"],
        ["Snapshots", "PnL/Delta historical log for active trades.", "Make", "READY"],
        ["Import Queue", "New trade detection before final write.", "Make", "READY"],
        ["AI Export", "Prompt-ready trade review layer.", "Joao", "READY"],
        ["Cristiano View", "Dashboard and filtered trade views only.", "Joao", "NEXT"],
    ]
    for row_idx, values in enumerate(rows, start=6):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)
    data_cell_style(ws, 6, 20, 4)

    section_label(ws, 13, "Migration Checklist")
    checklist = [
        "Map current monthly trade blocks into db_trade_registry",
        "Map scraper output into db_trade_legs and db_trade_registry",
        "Redirect Make snapshot sync into db_snapshots",
        "Retire direct reads from visual tabs",
        "Publish Cristiano shared copy",
    ]
    for idx, item in enumerate(checklist, start=14):
        ws[f"A{idx}"] = "[ ]"
        ws[f"B{idx}"] = item
        ws[f"B{idx}"].alignment = Alignment(wrap_text=True)


def add_validations(wb: Workbook) -> None:
    registry = wb["db_trade_registry"]
    ai_export = wb["AI Export"]

    status_dv = DataValidation(type="list", formula1="=dim_lookups!$A$6:$A$11", allow_blank=True)
    book_dv = DataValidation(type="list", formula1="=dim_lookups!$C$6:$C$9", allow_blank=True)
    owner_dv = DataValidation(type="list", formula1="=dim_lookups!$E$6:$E$8", allow_blank=True)
    sleeve_dv = DataValidation(type="list", formula1="=dim_lookups!$G$6:$G$9", allow_blank=True)
    underlying_dv = DataValidation(type="list", formula1="=dim_lookups!$I$6:$I$13", allow_blank=True)
    setup_dv = DataValidation(type="list", formula1="=dim_lookups!$K$6:$K$11", allow_blank=True)
    ai_ready_dv = DataValidation(type="list", formula1="=dim_lookups!$Q$6:$Q$7", allow_blank=True)
    trade_picker_dv = DataValidation(type="list", formula1="=db_trade_registry!$A$6:$A$5000", allow_blank=True)

    for dv in [status_dv, book_dv, owner_dv, sleeve_dv, underlying_dv, setup_dv, ai_ready_dv]:
        registry.add_data_validation(dv)
    ai_export.add_data_validation(trade_picker_dv)

    status_dv.add("E6:E5000")
    book_dv.add("B6:B5000")
    owner_dv.add("C6:C5000")
    sleeve_dv.add("D6:D5000")
    underlying_dv.add("I6:I5000")
    setup_dv.add("G6:G5000")
    ai_ready_dv.add("AD6:AD5000")
    trade_picker_dv.add("B6")


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    visible = [
        "Dashboard", "Trade Monitor", "Active Trades", "Closed Trades",
        "Performance", "AI Export", "Playbook", "Ops Control",
    ]
    hidden = [
        "db_trade_registry", "db_trade_legs", "db_snapshots", "db_events",
        "db_import_queue", "db_make_control", "dim_lookups",
    ]
    for name in visible + hidden:
        wb.create_sheet(title=name)

    build_dashboard(wb["Dashboard"])
    build_monitor(wb["Trade Monitor"])
    build_active(wb["Active Trades"])
    build_closed(wb["Closed Trades"])
    build_performance(wb["Performance"])
    build_ai_export(wb["AI Export"])
    build_playbook(wb["Playbook"])
    build_ops_control(wb["Ops Control"])

    build_registry(wb["db_trade_registry"])
    build_trade_legs(wb["db_trade_legs"])
    build_snapshots(wb["db_snapshots"])
    build_events(wb["db_events"])
    build_import_queue(wb["db_import_queue"])
    build_make_control(wb["db_make_control"])
    build_lookups(wb["dim_lookups"])

    add_validations(wb)

    for name in hidden:
        wb[name].sheet_state = "hidden"

    wb.properties.creator = "OpenAI Codex"
    wb.properties.title = "OP Control Panel Ideal Template"
    wb.properties.subject = "Redesigned trading control panel for Google Sheets"

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Workbook generated: {OUTPUT}")


if __name__ == "__main__":
    main()
