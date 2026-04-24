#!/usr/bin/env python3
"""
clean_db_robots.py
------------------
Remove as ~51.000 linhas de placeholder (vazias) do db_robots e salva
uma cópia limpa do workbook. A cópia pode ser re-uploadada ao Google Sheets.

Uso:
    python scripts/clean_db_robots.py

Output:
    Data/OP Control Panel - Clean.xlsx

IMPORTANTE: Não sobrescreve o arquivo original. Verifique a cópia limpa
antes de substituir no Google Drive.
"""

import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT      = Path(__file__).parent.parent
ORIG      = ROOT / "data" / "control_panel" / "OP Control Panel.xlsx"
CLEAN_OUT = ROOT / "data" / "control_panel" / "OP Control Panel - Clean.xlsx"


def clean_db_robots(xlsx_path: Path, output_path: Path) -> dict:
    print(f"[clean] Carregando {xlsx_path.name} (pode demorar alguns segundos)...")
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["db_robots"]

    total_rows = ws.max_row - 1  # excluindo header

    # Identificar linhas com dados reais
    rows_with_data = []
    rows_placeholder = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        env = row[1] if len(row) > 1 else None
        if env and str(env).strip():
            rows_with_data.append(row)
        else:
            rows_placeholder += 1

    print(f"[clean] Total de linhas: {total_rows:,}")
    print(f"[clean] Linhas com dados reais: {len(rows_with_data):,}")
    print(f"[clean] Linhas placeholder removidas: {rows_placeholder:,}")

    # Reescrever db_robots sem os placeholders
    # Deletar todas as linhas de dados (manter apenas header na row 1)
    # openpyxl não tem delete_rows eficiente para sheets grandes,
    # então reescrevemos o conteúdo
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    # Limpar toda a sheet abaixo do header
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).value = None

    # Reescrever dados reais
    for r_idx, row_data in enumerate(rows_with_data, start=2):
        for c_idx, val in enumerate(row_data, start=1):
            if c_idx <= ws.max_column:
                ws.cell(row=r_idx, column=c_idx).value = val

    print(f"[clean] Salvando {output_path.name}...")
    wb.save(output_path)

    return {
        "total_original": total_rows,
        "rows_kept":      len(rows_with_data),
        "rows_removed":   rows_placeholder,
        "output":         str(output_path),
    }


if __name__ == "__main__":
    if not ORIG.exists():
        print(f"[ERROR] Arquivo nao encontrado: {ORIG}")
        sys.exit(1)

    if CLEAN_OUT.exists():
        print(f"[AVISO] Arquivo de output ja existe: {CLEAN_OUT.name}")
        resp = input("Sobrescrever? [s/N] ").strip().lower()
        if resp != "s":
            print("[clean] Cancelado.")
            sys.exit(0)

    result = clean_db_robots(ORIG, CLEAN_OUT)

    print()
    print("=" * 50)
    print("  LIMPEZA CONCLUIDA")
    print("=" * 50)
    print(f"  Arquivo original:  {ORIG.name}")
    print(f"  Arquivo limpo:     {CLEAN_OUT.name}")
    print(f"  Linhas removidas:  {result['rows_removed']:,} placeholders")
    print(f"  Linhas mantidas:   {result['rows_kept']:,} registros reais")
    print()
    print("  PROXIMO PASSO:")
    print("  1. Abra 'OP Control Panel - Clean.xlsx' e verifique o db_robots")
    print("  2. Se correto, faca upload para o Google Drive como NOVA versao")
    print("  3. Atualize o Make Scenario 2 se necessario (proxima linha livre)")
    print("=" * 50)
